"""
BioFace — Moliya va Ish Haqi Hisobi Moduli
============================================
O'zbekiston davlat standartlariga muvofiq ish haqi hisoblash va dinamik dam olish kunlari.

Dam olish kunlari va bayramlar:
  - Dam olish kunlari (weekends) va bayramlar (holidays) https://bioface.uz/shifts (Shifts/Smenalar va Tavim)
    sahifasida kiritilgan dynamic Holiday ma'lumotlariga asosan aniqlanadi.
  - Odatiy yakshanba kunlari standart dam olish kuni sifatida hisoblanadi.

Soliq me'yorlari (2024-2025):
  - Jismoniy shaxslardan olinadigan daromad solig'i (JSODS/JSHIR):
      * Oylik eng kam ish haqi (EKISH) 2025: 1 050 000 UZS
      * Soliq chegirmasi: 1 EKISH miqdoridagi daromad soliqsiz
      * Qolgan qism: 12% soliq
  - Pensiya jamg'armasi (xodim ulushi): 8%
  - Ijtimoiy sug'urta (ish beruvchi tomonidan to'lanadi)

Hisoblash formulasi (davlat standarti):
  1. Brutto ish haqi = Asosiy maosh - Kechikish chegirilmalari - Yo'qlik chegirilmalari
  2. Pensiya ajratma = Brutto × 8% (xodim ulushi)
  3. Soliq bazasi = Brutto - Pensiya ajratma
  4. Soliq chegirilmasi = 1 × EKISH (chegirma)
  5. JSHIR = max(0, Soliq bazasi - Soliq chegirilmasi) × 12%
  6. Netto (qo'lga tegadigan) = Brutto - Pensiya ajratma - JSHIR
"""

from __future__ import annotations

import io
import math
import os
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, and_
from sqlalchemy.orm import Session, joinedload

from database import get_db
from models import (
    AccountTransfer,
    AttendanceLog,
    CashflowTransaction,
    Employee,
    FinanceAccount,
    Holiday,
)
from routers.dashboard import _resolve_allowed_org_ids
from utils.schedule_utils import (
    get_expected_end_dt,
    get_expected_start_dt,
    get_late_minutes,
    load_holiday_details,
    is_day_off_for_emp,
)
from utils.time_utils import now_tashkent

router = APIRouter()

# ─── O'ZBEKISTON SOLIQ KONSTANTALARI 2025 ────────────────────────────────────
# Eng kam ish haqi (EKISH) 2025 yil — 1 050 000 UZS
MINIMUM_WAGE_UZS: int = 1_050_000

# Jismoniy shaxslardan daromad solig'i (JSHIR) stavkasi: 12%
INCOME_TAX_RATE: Decimal = Decimal("0.12")

# Pensiya jamg'armasi (xodim ulushi): 8%
PENSION_RATE: Decimal = Decimal("0.08")

# Soliqdan ozod qilinadigan miqdor (1 × EKISH)
TAX_FREE_AMOUNT: int = MINIMUM_WAGE_UZS

# ─── YORDAMCHI FUNKSIYALAR ───────────────────────────────────────────────────


def _to_int_uzs(value: Any) -> int:
    """Har qanday miqdorni butun UZS ga aylantiradi (matematika bo'yicha yaxlitlash)."""
    try:
        d = Decimal(str(value or 0))
        return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0


def _get_active_status(status_records: list, day_dt: date):
    """Berilgan kun uchun faol xodim status yozuvini qaytaradi (vacation, sick_leave, etc.)."""
    for r in status_records:
        if r.start_date <= day_dt and (r.end_date is None or r.end_date >= day_dt):
            return r
    return None


def load_employee_statuses(
    db: Session,
    employee_ids: list[int],
    start_date: date,
    end_date: date,
) -> dict[int, list]:
    """Xodimlar holat yozuvlarini yuklaydi."""
    from models import EmployeeStatusRecord

    records = (
        db.query(EmployeeStatusRecord)
        .filter(
            EmployeeStatusRecord.employee_id.in_(employee_ids),
            EmployeeStatusRecord.start_date <= end_date,
        )
        .all()
    )
    status_by_emp: dict[int, list] = defaultdict(list)
    for r in records:
        end_ok = r.end_date is None or r.end_date >= start_date
        if end_ok:
            status_by_emp[r.employee_id].append(r)
    return status_by_emp


# ─── ISH HAQI HISOBLASH YADROSI ──────────────────────────────────────────────


def compute_salary_breakdown(
    base_salary: int,
    working_days: int,
    absent_days: int,
    late_details: list[tuple[int, int]],
    apply_tax: bool = True,
) -> dict[str, int]:
    """
    O'zbekiston mehnat va soliq qonunchiligiga muvofiq ish haqi hisobi.
    """
    if working_days <= 0 or base_salary <= 0:
        return {
            "base": base_salary,
            "gross": 0,
            "absent_deduction": 0,
            "late_deduction": 0,
            "total_deduction": 0,
            "pension_deduction": 0,
            "income_tax": 0,
            "net": 0,
            "working_days": working_days,
            "absent_days": absent_days,
        }

    daily_rate = Decimal(str(base_salary)) / Decimal(str(working_days))

    # 1. Yo'qlik chegirilmasi
    absent_deduction = _to_int_uzs(daily_rate * Decimal(str(absent_days)))

    # 2. Kechikish chegirilmasi
    late_deduction_total = Decimal("0")
    for late_mins, work_mins in late_details:
        if work_mins > 0 and late_mins > 0:
            minutely_rate = daily_rate / Decimal(str(work_mins))
            late_deduction_total += minutely_rate * Decimal(str(late_mins))
    late_deduction = _to_int_uzs(late_deduction_total)

    total_deduction = absent_deduction + late_deduction
    gross = max(0, base_salary - total_deduction)

    if not apply_tax or gross == 0:
        return {
            "base": base_salary,
            "gross": gross,
            "absent_deduction": absent_deduction,
            "late_deduction": late_deduction,
            "total_deduction": total_deduction,
            "pension_deduction": 0,
            "income_tax": 0,
            "net": gross,
            "working_days": working_days,
            "absent_days": absent_days,
        }

    # 3. Pensiya ajratma (xodim ulushi 8%)
    gross_d = Decimal(str(gross))
    pension_deduction = _to_int_uzs(gross_d * PENSION_RATE)

    # 4. Soliq bazasi = Brutto - Pensiya ajratma
    tax_base = gross_d - Decimal(str(pension_deduction))

    # 5. Soliqdan ozod qilinadigan miqdor (1 × EKISH)
    tax_free = Decimal(str(TAX_FREE_AMOUNT))

    # 6. JSHIR (12%)
    taxable_income = max(Decimal("0"), tax_base - tax_free)
    income_tax = _to_int_uzs(taxable_income * INCOME_TAX_RATE)

    # 7. Netto (Qo'lga tegadigan)
    net = max(0, gross - pension_deduction - income_tax)

    return {
        "base": base_salary,
        "gross": gross,
        "absent_deduction": absent_deduction,
        "late_deduction": late_deduction,
        "total_deduction": total_deduction,
        "pension_deduction": pension_deduction,
        "income_tax": income_tax,
        "net": net,
        "working_days": working_days,
        "absent_days": absent_days,
    }


def _build_emp_attendance_summary(
    emp: Employee,
    logs_by_emp_day: dict,
    status_by_emp: dict,
    holiday_details: dict,
    target_year: int,
    target_month: int,
    days_in_month: int,
    today_date: date,
) -> dict[str, Any]:
    """
    Xodim uchun bir oylik davomat xulosasini DINAMIK hisoblaydi.
    Dam olish kunlari https://bioface.uz/shifts jadvalidan yuklanadi.
    """
    working_days = 0
    absent_days = 0
    late_details: list[tuple[int, int]] = []
    late_count = 0

    for day_num in range(1, days_in_month + 1):
        day_dt = date(target_year, target_month, day_num)
        day_str = day_dt.isoformat()

        # Dinamik dam olish / bayram tekshiruvi (Shifts jadvali bo'yicha)
        is_off, _ = is_day_off_for_emp(day_dt, emp.organization_id, holiday_details)
        if is_off:
            continue

        active_status = _get_active_status(status_by_emp.get(emp.id, []), day_dt)

        if active_status and active_status.status_type in ("resigned", "suspended"):
            continue

        working_days += 1

        if day_dt > today_date:
            continue

        is_excused = active_status and active_status.status_type in (
            "vacation",
            "sick_leave",
            "business_trip",
            "other",
        )

        seen_pair = logs_by_emp_day.get((emp.id, day_str))
        is_absent = False

        if not seen_pair:
            is_absent = True
        else:
            first_seen = seen_pair[0]
            expected_end = get_expected_end_dt(emp, day_dt)
            if first_seen >= expected_end:
                is_absent = True

        if is_absent and not is_excused:
            absent_days += 1
        elif not is_absent and seen_pair:
            first_seen = seen_pair[0]
            late_mins = get_late_minutes(emp, day_dt, first_seen)
            if late_mins > 0:
                expected_start = get_expected_start_dt(emp, day_dt)
                expected_end = get_expected_end_dt(emp, day_dt)
                work_mins = max(
                    60,
                    int((expected_end - expected_start).total_seconds() // 60),
                )
                late_details.append((late_mins, work_mins))
                late_count += 1

    return {
        "working_days": working_days,
        "absent_days": absent_days,
        "late_details": late_details,
        "late_count": late_count,
    }


def _get_employee_base_query(db: Session, allowed_orgs: list[int], month_start: datetime, month_end: datetime):
    """Xodimlar uchun asosiy query."""
    from models import EmployeeStatusRecord

    active_emp_ids = (
        db.query(AttendanceLog.employee_id)
        .filter(
            AttendanceLog.timestamp >= month_start,
            AttendanceLog.timestamp < month_end,
        )
        .distinct()
    )
    status_emp_ids = (
        db.query(EmployeeStatusRecord.employee_id)
        .filter(
            EmployeeStatusRecord.start_date <= month_end.date(),
            or_(
                EmployeeStatusRecord.end_date == None,
                EmployeeStatusRecord.end_date >= month_start.date(),
            ),
        )
        .distinct()
    )
    return (
        db.query(Employee)
        .options(
            joinedload(Employee.organization),
            joinedload(Employee.schedule),
            joinedload(Employee.position_ref),
        )
        .filter(
            or_(
                Employee.has_access == True,
                Employee.id.in_(active_emp_ids),
                Employee.id.in_(status_emp_ids),
            )
        )
        .filter(Employee.organization_id.in_(allowed_orgs))
    )


# ─── API ENDPOINTLAR ──────────────────────────────────────────────────────────


@router.get("/api/finance/salaries")
def get_salaries(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query("all"),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    apply_tax: bool = Query(True, description="Soliq va pensiya chegirilmalarini qo'llash"),
    db: Session = Depends(get_db),
):
    """
    Xodimlar ish haqi ro'yxati.
    Dam olish kunlari https://bioface.uz/shifts (Shifts va Bayramlar) bo'yicha dinamik aniqlanadi.
    """
    _empty_stats = {
        "totalBase": 0,
        "totalGross": 0,
        "totalDeductions": 0,
        "totalPension": 0,
        "totalTax": 0,
        "totalNet": 0,
        "paidSum": 0,
        "unpaidSum": 0,
        "minWage": MINIMUM_WAGE_UZS,
        "taxRate": float(INCOME_TAX_RATE * 100),
        "pensionRate": float(PENSION_RATE * 100),
    }

    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        return {"salaries": [], "stats": _empty_stats}

    now = now_tashkent()
    target_year = year or now.year
    target_month = month or now.month

    if not (1 <= target_month <= 12):
        raise HTTPException(status_code=400, detail="Oy 1-12 orasida bo'lishi kerak")

    days_in_month = monthrange(target_year, target_month)[1]
    month_start = datetime(target_year, target_month, 1)
    month_end = (
        datetime(target_year + 1, 1, 1) if target_month == 12
        else datetime(target_year, target_month + 1, 1)
    )

    query = _get_employee_base_query(db, allowed_orgs, month_start, month_end)

    if organization_id is not None:
        if organization_id not in allowed_orgs:
            return {"salaries": [], "stats": _empty_stats}
        query = query.filter(Employee.organization_id == organization_id)

    if branch_id is not None:
        query = query.filter(Employee.branch_id == branch_id)

    if search:
        s = f"%{search.strip().lower()}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(s),
                Employee.last_name.ilike(s),
                Employee.middle_name.ilike(s),
                Employee.position.ilike(s),
            )
        )

    employees = query.all()
    if not employees:
        return {"salaries": [], "stats": _empty_stats}

    emp_ids = [emp.id for emp in employees]

    logs = (
        db.query(AttendanceLog)
        .filter(
            AttendanceLog.employee_id.in_(emp_ids),
            AttendanceLog.timestamp >= month_start,
            AttendanceLog.timestamp < month_end,
        )
        .order_by(AttendanceLog.employee_id, AttendanceLog.timestamp.asc())
        .all()
    )

    logs_by_emp_day: dict[tuple, list] = {}
    for log in logs:
        if not log.timestamp:
            continue
        day_str = log.timestamp.strftime("%Y-%m-%d")
        key = (log.employee_id, day_str)
        if key not in logs_by_emp_day:
            logs_by_emp_day[key] = [log.timestamp, log.timestamp]
        else:
            if log.timestamp < logs_by_emp_day[key][0]:
                logs_by_emp_day[key][0] = log.timestamp
            if log.timestamp > logs_by_emp_day[key][1]:
                logs_by_emp_day[key][1] = log.timestamp

    # Dinamik bayram va dam olish kunlarini yuklash (/shifts bo'yicha)
    holiday_details = load_holiday_details(
        db,
        start_date=month_start.date(),
        end_date=date(target_year, target_month, days_in_month),
        organization_ids=allowed_orgs,
    )

    today_date = now.date()
    status_by_emp = load_employee_statuses(
        db, emp_ids, month_start.date(), date(target_year, target_month, days_in_month)
    )

    salary_list = []
    total_base = 0
    total_gross = 0
    total_deductions = 0
    total_pension = 0
    total_tax = 0
    total_net = 0
    paid_sum = 0
    unpaid_sum = 0

    for emp in employees:
        base_salary = _to_int_uzs(emp.salary or 0)

        att = _build_emp_attendance_summary(
            emp=emp,
            logs_by_emp_day=logs_by_emp_day,
            status_by_emp=status_by_emp,
            holiday_details=holiday_details,
            target_year=target_year,
            target_month=target_month,
            days_in_month=days_in_month,
            today_date=today_date,
        )

        bd = compute_salary_breakdown(
            base_salary=base_salary,
            working_days=att["working_days"],
            absent_days=att["absent_days"],
            late_details=att["late_details"],
            apply_tax=apply_tax,
        )

        emp_status = emp.salary_status or "unpaid"

        if status != "all":
            if status == "unpaid":
                if emp_status not in ("unpaid", "advance"):
                    continue
            elif emp_status != status:
                continue

        full_name = " ".join(
            p for p in [emp.last_name, emp.first_name, emp.middle_name] if p and str(p).strip()
        ).strip()
        role_name = (
            emp.position_ref.name if emp.position_ref else (emp.position or emp.employee_type or "Xodim")
        )

        below_min_wage = base_salary > 0 and base_salary < MINIMUM_WAGE_UZS

        salary_list.append({
            "id": emp.id,
            "uuid": emp.uuid,
            "name": full_name,
            "role": role_name,
            "base": bd["base"],
            "gross": bd["gross"],
            "lateCount": att["late_count"],
            "lateDeduction": bd["late_deduction"],
            "absentCount": att["absent_days"],
            "absentDeduction": bd["absent_deduction"],
            "totalDeduction": bd["total_deduction"],
            "pensionDeduction": bd["pension_deduction"],
            "incomeTax": bd["income_tax"],
            "net": bd["net"],
            "workingDays": att["working_days"],
            "status": emp_status,
            "organization_id": emp.organization_id,
            "branch_id": emp.branch_id,
            "belowMinWage": below_min_wage,
            "finalAmount": bd["net"],
        })

        total_base += bd["base"]
        total_gross += bd["gross"]
        total_deductions += bd["total_deduction"]
        total_pension += bd["pension_deduction"]
        total_tax += bd["income_tax"]
        total_net += bd["net"]

        if emp_status == "paid":
            paid_sum += bd["net"]
        elif emp_status == "advance":
            paid_sum += bd["net"] // 2
            unpaid_sum += bd["net"] - bd["net"] // 2
        else:
            unpaid_sum += bd["net"]

    return {
        "salaries": salary_list,
        "stats": {
            "totalBase": total_base,
            "totalGross": total_gross,
            "totalDeductions": total_deductions,
            "totalPension": total_pension,
            "totalTax": total_tax,
            "totalNet": total_net,
            "paidSum": paid_sum,
            "unpaidSum": unpaid_sum,
            "minWage": MINIMUM_WAGE_UZS,
            "taxRate": float(INCOME_TAX_RATE * 100),
            "pensionRate": float(PENSION_RATE * 100),
        },
    }


@router.post("/api/finance/salaries/{emp_id}/pay")
def pay_salary(
    request: Request,
    emp_id: str,
    pay_type: str = Query("full"),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    apply_tax: bool = Query(True),
    db: Session = Depends(get_db),
):
    """Xodimga ish haqi to'lovi."""
    allowed_orgs = _resolve_allowed_org_ids(request, db)

    if emp_id.isdigit():
        emp = db.query(Employee).options(
            joinedload(Employee.organization),
            joinedload(Employee.schedule),
        ).filter(Employee.id == int(emp_id)).first()
    else:
        emp = db.query(Employee).options(
            joinedload(Employee.organization),
            joinedload(Employee.schedule),
        ).filter(Employee.uuid == emp_id).first()

    if not emp:
        raise HTTPException(status_code=404, detail="Xodim topilmadi")
    if emp.organization_id not in allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    now = now_tashkent()
    target_year = year or now.year
    target_month = month or now.month

    days_in_month = monthrange(target_year, target_month)[1]
    month_start = datetime(target_year, target_month, 1)
    month_end = (
        datetime(target_year + 1, 1, 1) if target_month == 12
        else datetime(target_year, target_month + 1, 1)
    )

    logs = (
        db.query(AttendanceLog)
        .filter(
            AttendanceLog.employee_id == emp.id,
            AttendanceLog.timestamp >= month_start,
            AttendanceLog.timestamp < month_end,
        )
        .order_by(AttendanceLog.timestamp.asc())
        .all()
    )

    logs_by_emp_day: dict[tuple, list] = {}
    for log in logs:
        if not log.timestamp:
            continue
        day_str = log.timestamp.strftime("%Y-%m-%d")
        key = (emp.id, day_str)
        if key not in logs_by_emp_day:
            logs_by_emp_day[key] = [log.timestamp, log.timestamp]
        else:
            if log.timestamp < logs_by_emp_day[key][0]:
                logs_by_emp_day[key][0] = log.timestamp
            if log.timestamp > logs_by_emp_day[key][1]:
                logs_by_emp_day[key][1] = log.timestamp

    holiday_details = load_holiday_details(
        db,
        start_date=month_start.date(),
        end_date=date(target_year, target_month, days_in_month),
        organization_ids=[emp.organization_id],
    )

    status_by_emp = load_employee_statuses(
        db, [emp.id], month_start.date(), date(target_year, target_month, days_in_month)
    )

    att = _build_emp_attendance_summary(
        emp=emp,
        logs_by_emp_day=logs_by_emp_day,
        status_by_emp=status_by_emp,
        holiday_details=holiday_details,
        target_year=target_year,
        target_month=target_month,
        days_in_month=days_in_month,
        today_date=now.date(),
    )

    base_salary = _to_int_uzs(emp.salary or 0)
    bd = compute_salary_breakdown(
        base_salary=base_salary,
        working_days=att["working_days"],
        absent_days=att["absent_days"],
        late_details=att["late_details"],
        apply_tax=apply_tax,
    )

    amount_to_pay = bd["net"] // 2 if pay_type == "advance" else bd["net"]

    emp.salary_status = "advance" if pay_type == "advance" else "paid"

    full_name = " ".join(
        p for p in [emp.last_name, emp.first_name, emp.middle_name] if p and str(p).strip()
    ).strip()

    acc = (
        db.query(FinanceAccount)
        .filter(
            FinanceAccount.organization_id == emp.organization_id,
            FinanceAccount.type == "bank",
        )
        .first()
    ) or db.query(FinanceAccount).filter(FinanceAccount.organization_id == emp.organization_id).first()

    acc_id = None
    if acc:
        if acc.balance < amount_to_pay:
            raise HTTPException(
                status_code=400,
                detail=f"Hisobda mablag' yetarli emas. Mavjud: {int(acc.balance):,} UZS, kerak: {amount_to_pay:,} UZS",
            )
        acc.balance -= float(amount_to_pay)
        acc_id = acc.id

    pay_label = "Avans" if pay_type == "advance" else "To'liq"
    month_names = [
        "", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentyabr", "Oktyabr", "Noyabr", "Dekabr",
    ]
    month_name = month_names[target_month]

    tx = CashflowTransaction(
        description=f"Ish haqi — {full_name} ({pay_label})",
        type="expense",
        amount=float(amount_to_pay),
        comment=(
            f"{month_name} {target_year} | Brutto: {bd['gross']:,} | "
            f"Pensiya: {bd['pension_deduction']:,} | JSHIR: {bd['income_tax']:,} | "
            f"Netto: {bd['net']:,} UZS"
        ),
        date=now.strftime("%Y-%m-%d"),
        organization_id=emp.organization_id,
        branch_id=emp.branch_id,
        employee_id=emp.id,
        account_id=acc_id,
    )
    db.add(tx)
    db.commit()

    return {
        "status": "success",
        "message": "Oylik muvaffaqiyatli to'landi",
        "paid": amount_to_pay,
        "breakdown": bd,
    }


@router.post("/api/finance/salaries/reset")
def reset_salaries(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    query = db.query(Employee)
    if organization_id is not None:
        if organization_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="Ruxsat berilmagan")
        query = query.filter(Employee.organization_id == organization_id)
    else:
        query = query.filter(Employee.organization_id.in_(allowed_orgs))

    if branch_id is not None:
        query = query.filter(Employee.branch_id == branch_id)

    employees = query.all()
    for emp in employees:
        emp.salary_status = "unpaid"
    db.commit()
    return {"status": "success", "message": "Barcha to'lov holatlari yangilandi", "count": len(employees)}


@router.get("/api/finance/kpi")
def get_kpi_stats(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """
    KPI ko'rsatkichlari bo'yicha dinamik statistikalar.
    Dam olish kunlari https://bioface.uz/shifts jadvalidan olinadi.
    """
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        return {"kpis": []}

    now = now_tashkent()
    target_year = year or now.year
    target_month = month or now.month

    days_in_month = monthrange(target_year, target_month)[1]
    month_start = datetime(target_year, target_month, 1)
    month_end = (
        datetime(target_year + 1, 1, 1) if target_month == 12
        else datetime(target_year, target_month + 1, 1)
    )

    query = _get_employee_base_query(db, allowed_orgs, month_start, month_end)

    if organization_id is not None:
        if organization_id not in allowed_orgs:
            return {"kpis": []}
        query = query.filter(Employee.organization_id == organization_id)

    if branch_id is not None:
        query = query.filter(Employee.branch_id == branch_id)

    if search:
        s = f"%{search.strip().lower()}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(s),
                Employee.last_name.ilike(s),
                Employee.middle_name.ilike(s),
                Employee.position.ilike(s),
            )
        )

    employees = query.all()
    if not employees:
        return {"kpis": []}

    emp_ids = [emp.id for emp in employees]

    logs = (
        db.query(AttendanceLog)
        .filter(
            AttendanceLog.employee_id.in_(emp_ids),
            AttendanceLog.timestamp >= month_start,
            AttendanceLog.timestamp < month_end,
        )
        .order_by(AttendanceLog.employee_id, AttendanceLog.timestamp.asc())
        .all()
    )

    logs_by_emp_day: dict[tuple, list] = {}
    for log in logs:
        if not log.timestamp:
            continue
        day_str = log.timestamp.strftime("%Y-%m-%d")
        key = (log.employee_id, day_str)
        if key not in logs_by_emp_day:
            logs_by_emp_day[key] = [log.timestamp, log.timestamp]
        else:
            if log.timestamp < logs_by_emp_day[key][0]:
                logs_by_emp_day[key][0] = log.timestamp
            if log.timestamp > logs_by_emp_day[key][1]:
                logs_by_emp_day[key][1] = log.timestamp

    holiday_details = load_holiday_details(
        db,
        start_date=month_start.date(),
        end_date=date(target_year, target_month, days_in_month),
        organization_ids=allowed_orgs,
    )

    today_date = now.date()
    status_by_emp = load_employee_statuses(
        db, emp_ids, month_start.date(), date(target_year, target_month, days_in_month)
    )

    kpi_list = []
    for emp in employees:
        att = _build_emp_attendance_summary(
            emp=emp,
            logs_by_emp_day=logs_by_emp_day,
            status_by_emp=status_by_emp,
            holiday_details=holiday_details,
            target_year=target_year,
            target_month=target_month,
            days_in_month=days_in_month,
            today_date=today_date,
        )

        base_salary = _to_int_uzs(emp.salary or 0)
        total_overtime_hours = Decimal("0")
        ontime_days = 0

        for day_num in range(1, days_in_month + 1):
            day_dt = date(target_year, target_month, day_num)
            day_str = day_dt.isoformat()

            is_off, _ = is_day_off_for_emp(day_dt, emp.organization_id, holiday_details)
            if is_off or day_dt > today_date:
                continue

            seen_pair = logs_by_emp_day.get((emp.id, day_str))
            if not seen_pair:
                continue

            first_seen, last_seen = seen_pair[0], seen_pair[1]
            expected_start = get_expected_start_dt(emp, day_dt)
            expected_end = get_expected_end_dt(emp, day_dt)

            expected_duration_sec = (expected_end - expected_start).total_seconds()
            if expected_duration_sec <= 0:
                continue

            worked_sec = (last_seen - first_seen).total_seconds()
            overtime_sec = max(0.0, worked_sec - expected_duration_sec)
            total_overtime_hours += Decimal(str(overtime_sec / 3600.0))

            late_mins = get_late_minutes(emp, day_dt, first_seen)
            if late_mins == 0:
                ontime_days += 1

        overtime_bonus = 0
        if att["working_days"] > 0 and base_salary > 0 and total_overtime_hours > 0:
            hourly_rate = Decimal(str(base_salary)) / (
                Decimal(str(att["working_days"])) * Decimal("8")
            )
            overtime_bonus = _to_int_uzs(hourly_rate * Decimal("1.5") * total_overtime_hours)

        working_days_so_far = sum(
            1
            for d in range(1, min(today_date.day + 1, days_in_month + 1))
            if not is_day_off_for_emp(date(target_year, target_month, d), emp.organization_id, holiday_details)[0]
        )
        present_days = max(0, working_days_so_far - att["absent_days"])

        attendance_rate = (
            int(round((present_days / working_days_so_far) * 100))
            if working_days_so_far > 0
            else 100
        )
        ontime_rate = (
            int(round((ontime_days / present_days) * 100)) if present_days > 0 else 100
        )
        score = int(round(attendance_rate * 0.6 + ontime_rate * 0.4))

        dept_name = (
            emp.department_ref.name if emp.department_ref else (emp.department or emp.employee_type or "Xodim")
        )

        kpi_list.append({
            "id": emp.id,
            "uuid": emp.uuid,
            "name": f"{emp.last_name or ''} {emp.first_name or ''}".strip(),
            "dept": dept_name,
            "attendance": attendance_rate,
            "ontime": ontime_rate,
            "score": score,
            "baseSalary": base_salary,
            "overtimeBonus": overtime_bonus,
            "totalDeductions": att["absent_days"] + att["late_count"],
            "ontimeCount": ontime_days,
            "lateCount": att["late_count"],
            "absentCount": att["absent_days"],
            "workingDays": att["working_days"],
            "isAwarded": False,
        })

    return {"kpis": kpi_list}


@router.post("/api/finance/kpi/{emp_id}/reward")
def award_kpi(
    request: Request,
    emp_id: str,
    payload: dict,
    db: Session = Depends(get_db),
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    if emp_id.isdigit():
        emp = db.query(Employee).filter(Employee.id == int(emp_id)).first()
    else:
        emp = db.query(Employee).filter(Employee.uuid == emp_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Xodim topilmadi")
    if emp.organization_id not in allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    amount = _to_int_uzs(payload.get("amount", 0))
    comment = str(payload.get("comment", "") or "")

    if amount <= 0:
        raise HTTPException(status_code=400, detail="Mukofot miqdori musbat bo'lishi kerak")

    full_name = " ".join(
        p for p in [emp.last_name, emp.first_name, emp.middle_name] if p and str(p).strip()
    ).strip()

    acc = (
        db.query(FinanceAccount)
        .filter(FinanceAccount.organization_id == emp.organization_id, FinanceAccount.type == "bank")
        .first()
    ) or db.query(FinanceAccount).filter(FinanceAccount.organization_id == emp.organization_id).first()

    acc_id = None
    if acc:
        if acc.balance < amount:
            raise HTTPException(
                status_code=400,
                detail=f"Hisobda mablag' yetarli emas. Mavjud: {int(acc.balance):,} UZS",
            )
        acc.balance -= float(amount)
        acc_id = acc.id

    tx = CashflowTransaction(
        description=f"KPI mukofot — {full_name}",
        type="expense",
        amount=float(amount),
        comment=comment,
        date=now_tashkent().strftime("%Y-%m-%d"),
        organization_id=emp.organization_id,
        branch_id=emp.branch_id,
        employee_id=emp.id,
        account_id=acc_id,
    )
    db.add(tx)
    db.commit()
    return {"status": "success", "message": "Mukofot muvaffaqiyatli saqlandi", "amount": amount}


@router.get("/api/finance/tax-info")
def get_tax_info(request: Request):
    """Soliq me'yorlari ma'lumotlari."""
    return {
        "country": "UZ",
        "year": now_tashkent().year,
        "minimum_wage": MINIMUM_WAGE_UZS,
        "income_tax_rate": float(INCOME_TAX_RATE),
        "income_tax_rate_pct": float(INCOME_TAX_RATE * 100),
        "pension_rate": float(PENSION_RATE),
        "pension_rate_pct": float(PENSION_RATE * 100),
        "tax_free_amount": TAX_FREE_AMOUNT,
    }


# ─── CASHFLOW ─────────────────────────────────────────────────────────────────


@router.get("/api/finance/cashflow")
def get_cashflow(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    type: Optional[str] = Query("all"),
    db: Session = Depends(get_db),
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        return {"transactions": [], "stats": {"income": 0, "expense": 0, "profit": 0}}

    query = db.query(CashflowTransaction)

    if organization_id is not None:
        if organization_id not in allowed_orgs:
            return {"transactions": [], "stats": {"income": 0, "expense": 0, "profit": 0}}
        query = query.filter(CashflowTransaction.organization_id == organization_id)
    else:
        query = query.filter(CashflowTransaction.organization_id.in_(allowed_orgs))

    if branch_id is not None:
        query = query.filter(CashflowTransaction.branch_id == branch_id)

    if type != "all":
        query = query.filter(CashflowTransaction.type == type)

    if search:
        s = f"%{search.strip().lower()}%"
        query = query.outerjoin(Employee, CashflowTransaction.employee_id == Employee.id).filter(
            or_(
                CashflowTransaction.description.ilike(s),
                CashflowTransaction.comment.ilike(s),
                Employee.first_name.ilike(s),
                Employee.last_name.ilike(s),
            )
        )

    transactions_list = query.order_by(
        CashflowTransaction.date.desc(), CashflowTransaction.id.desc()
    ).all()

    income = Decimal("0")
    expense = Decimal("0")
    formatted_txs = []

    for tx in transactions_list:
        if tx.type == "income":
            income += Decimal(str(tx.amount or 0))
        else:
            expense += Decimal(str(tx.amount or 0))

        emp_name = None
        if tx.employee:
            emp_name = " ".join(
                p for p in [tx.employee.last_name, tx.employee.first_name] if p and str(p).strip()
            ).strip()

        formatted_txs.append({
            "id": tx.id,
            "uuid": tx.uuid,
            "desc": tx.description,
            "type": tx.type,
            "amount": float(tx.amount or 0),
            "comment": tx.comment,
            "date": tx.date,
            "employee_id": tx.employee_id,
            "employee_name": emp_name,
            "organization_id": tx.organization_id,
            "branch_id": tx.branch_id,
            "account_id": tx.account_id,
            "account_name": tx.account.name_uz if tx.account else None,
        })

    return {
        "transactions": formatted_txs,
        "stats": {
            "income": float(income),
            "expense": float(expense),
            "profit": float(income - expense),
        },
    }


@router.post("/api/finance/cashflow")
def add_cashflow_transaction(
    request: Request,
    payload: dict,
    db: Session = Depends(get_db),
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    description = str(payload.get("desc") or "").strip()
    tx_type = str(payload.get("type") or "").strip()
    amount_raw = payload.get("amount")
    comment = payload.get("comment")
    org_id = payload.get("organization_id")
    branch_id = payload.get("branch_id")
    date_val = payload.get("date")
    account_id = payload.get("account_id")

    if not description or not tx_type or amount_raw is None:
        raise HTTPException(status_code=400, detail="Tavsif, tur va miqdor kiritilishi shart")

    if tx_type not in ("income", "expense"):
        raise HTTPException(status_code=400, detail="Tur: 'income' yoki 'expense' bo'lishi kerak")

    amount = _to_int_uzs(amount_raw)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Miqdor musbat bo'lishi kerak")

    if not org_id:
        org_id = allowed_orgs[0]
    else:
        org_id = int(org_id)
        if org_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="Ruxsat berilmagan tashkilot")

    if not date_val:
        date_val = now_tashkent().strftime("%Y-%m-%d")

    acc = None
    if account_id:
        acc = db.query(FinanceAccount).filter(FinanceAccount.id == int(account_id)).first()
        if not acc or acc.organization_id not in allowed_orgs:
            raise HTTPException(status_code=400, detail="Yaroqsiz hisob")

    if not acc:
        acc = (
            db.query(FinanceAccount)
            .filter(FinanceAccount.organization_id == org_id, FinanceAccount.type == "bank")
            .first()
        )
        if not acc:
            acc = db.query(FinanceAccount).filter(FinanceAccount.organization_id == org_id).first()

    acc_id = None
    if acc:
        if tx_type == "expense" and acc.balance < amount:
            raise HTTPException(
                status_code=400,
                detail=f"Hisobda mablag' yetarli emas. Mavjud: {int(acc.balance):,} UZS",
            )
        acc_id = acc.id
        if tx_type == "income":
            acc.balance += float(amount)
        else:
            acc.balance -= float(amount)

    tx = CashflowTransaction(
        description=description,
        type=tx_type,
        amount=float(amount),
        comment=comment,
        date=date_val,
        organization_id=org_id,
        branch_id=int(branch_id) if branch_id else None,
        account_id=acc_id,
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)
    return {"status": "success", "transaction_id": tx.id}


# ─── HISOB-KITOB (ACCOUNTS) ───────────────────────────────────────────────────


@router.get("/api/finance/accounts")
def get_finance_accounts(request: Request, db: Session = Depends(get_db)):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        return {"accounts": [], "transfers": [], "totalBalance": 0}

    for org_id in allowed_orgs:
        acc_count = db.query(FinanceAccount).filter(FinanceAccount.organization_id == org_id).count()
        if acc_count == 0:
            defaults = [
                {"name_uz": "Asosiy Kassa (Naqd)", "name_ru": "Основная касса (Наличные)", "type": "cash"},
                {"name_uz": "Milliy Bank (Hisob raqam)", "name_ru": "Национальный Банк (Расч. счет)", "type": "bank"},
                {"name_uz": "Korporativ Karta (Uzcard/Humo)", "name_ru": "Корпоративная карта (Uzcard/Humo)", "type": "card"},
                {"name_uz": "Zaxira jamg'armasi", "name_ru": "Резервный фонд", "type": "reserve"},
            ]
            for d in defaults:
                db.add(FinanceAccount(
                    name_uz=d["name_uz"],
                    name_ru=d["name_ru"],
                    balance=0.0,
                    type=d["type"],
                    organization_id=org_id,
                ))
            db.commit()

    accounts = (
        db.query(FinanceAccount)
        .filter(FinanceAccount.organization_id.in_(allowed_orgs))
        .order_by(FinanceAccount.id.asc())
        .all()
    )

    transfers_list = (
        db.query(AccountTransfer)
        .join(FinanceAccount, AccountTransfer.from_account_id == FinanceAccount.id)
        .filter(FinanceAccount.organization_id.in_(allowed_orgs))
        .order_by(AccountTransfer.date.desc(), AccountTransfer.id.desc())
        .all()
    )

    return {
        "accounts": [
            {
                "id": a.id,
                "uuid": a.uuid,
                "nameUz": a.name_uz,
                "nameRu": a.name_ru,
                "accountNumber": a.account_number,
                "balance": float(a.balance),
                "type": a.type,
                "organization_id": a.organization_id,
                "branch_id": a.branch_id,
            }
            for a in accounts
        ],
        "transfers": [
            {
                "id": tx.id,
                "uuid": tx.uuid,
                "from": tx.from_account.name_uz,
                "to": tx.to_account.name_uz,
                "amount": float(tx.amount),
                "date": tx.date,
                "desc": tx.description or "",
            }
            for tx in transfers_list
        ],
        "totalBalance": sum(float(a.balance) for a in accounts),
    }


@router.post("/api/finance/accounts")
def create_finance_account(request: Request, payload: dict, db: Session = Depends(get_db)):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    name_uz = str(payload.get("nameUz") or "").strip()
    name_ru = str(payload.get("nameRu") or "").strip()
    if not name_uz or not name_ru:
        raise HTTPException(status_code=400, detail="Nomlar kiritilishi shart")

    acc_type = str(payload.get("type") or "cash")
    if acc_type not in ("cash", "bank", "card", "reserve"):
        raise HTTPException(status_code=400, detail="Hisob turi: cash, bank, card, reserve")

    org_id = payload.get("organization_id")
    if not org_id:
        org_id = allowed_orgs[0]
    else:
        org_id = int(org_id)
        if org_id not in allowed_orgs:
            raise HTTPException(status_code=403, detail="Ruxsat berilmagan tashkilot")

    acc = FinanceAccount(
        name_uz=name_uz,
        name_ru=name_ru,
        account_number=payload.get("accountNumber"),
        balance=float(_to_int_uzs(payload.get("balance", 0))),
        type=acc_type,
        organization_id=org_id,
        branch_id=int(payload["branch_id"]) if payload.get("branch_id") else None,
    )
    db.add(acc)
    db.commit()
    db.refresh(acc)
    return {"status": "success", "account_id": acc.id}


@router.put("/api/finance/accounts/{acc_id}")
def update_finance_account(
    request: Request, acc_id: int, payload: dict, db: Session = Depends(get_db)
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    acc = db.query(FinanceAccount).filter(FinanceAccount.id == acc_id).first()
    if not acc:
        raise HTTPException(status_code=404, detail="Hisob topilmadi")
    if acc.organization_id not in allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    if payload.get("nameUz"):
        acc.name_uz = str(payload["nameUz"]).strip()
    if payload.get("nameRu"):
        acc.name_ru = str(payload["nameRu"]).strip()
    if payload.get("accountNumber") is not None:
        acc.account_number = payload["accountNumber"]
    if payload.get("balance") is not None:
        acc.balance = float(_to_int_uzs(payload["balance"]))
    if payload.get("type") and payload["type"] in ("cash", "bank", "card", "reserve"):
        acc.type = payload["type"]

    db.commit()
    return {"status": "success", "message": "Hisob muvaffaqiyatli tahrirlandi"}


@router.delete("/api/finance/accounts/{acc_id}")
def delete_finance_account(
    request: Request, acc_id: int, db: Session = Depends(get_db)
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    acc = db.query(FinanceAccount).filter(FinanceAccount.id == acc_id).first()
    if not acc:
        raise HTTPException(status_code=404, detail="Hisob topilmadi")
    if acc.organization_id not in allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")
    if abs(acc.balance) > 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Hisobda qoldiq bor ({int(acc.balance):,} UZS). Avval mablag'ni o'tkazing.",
        )
    db.delete(acc)
    db.commit()
    return {"status": "success", "message": "Hisob o'chirildi"}


@router.post("/api/finance/accounts/transfer")
def transfer_funds(request: Request, payload: dict, db: Session = Depends(get_db)):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")

    from_id = payload.get("from_account_id")
    to_id = payload.get("to_account_id")
    amount = _to_int_uzs(payload.get("amount", 0))
    desc = str(payload.get("description") or "")
    date_val = payload.get("date") or now_tashkent().strftime("%Y-%m-%d")

    if not from_id or not to_id:
        raise HTTPException(status_code=400, detail="Hisob raqamlar ko'rsatilmagan")
    if amount <= 0:
        raise HTTPException(status_code=400, detail="O'tkazma miqdori musbat bo'lishi kerak")
    if int(from_id) == int(to_id):
        raise HTTPException(status_code=400, detail="Bir xil hisoblar orasida o'tkazma bo'lmaydi")

    from_acc = db.query(FinanceAccount).filter(FinanceAccount.id == int(from_id)).first()
    to_acc = db.query(FinanceAccount).filter(FinanceAccount.id == int(to_id)).first()

    if not from_acc or not to_acc:
        raise HTTPException(status_code=404, detail="Hisob topilmadi")
    if from_acc.organization_id not in allowed_orgs or to_acc.organization_id not in allowed_orgs:
        raise HTTPException(status_code=403, detail="Ruxsat berilmagan")
    if from_acc.balance < amount:
        raise HTTPException(
            status_code=400,
            detail=f"Hisobda mablag' yetarli emas. Mavjud: {int(from_acc.balance):,} UZS, kerak: {amount:,} UZS",
        )

    from_acc.balance -= float(amount)
    to_acc.balance += float(amount)

    tx = AccountTransfer(
        from_account_id=from_acc.id,
        to_account_id=to_acc.id,
        amount=float(amount),
        description=desc,
        date=date_val,
    )
    db.add(tx)
    db.commit()
    return {"status": "success", "message": "O'tkazma muvaffaqiyatli bajarildi", "amount": amount}


# ─── TABEL (DAVOMAT JADVALI) ──────────────────────────────────────────────────


def _fetch_tabel_data_internal(
    request: Request,
    organization_id: Optional[int],
    branch_id: Optional[int],
    department_id: Optional[int],
    position_id: Optional[int],
    search: Optional[str],
    year: Optional[int],
    month: Optional[int],
    db: Session,
):
    allowed_orgs = _resolve_allowed_org_ids(request, db)
    if not allowed_orgs:
        return [], 30, now_tashkent().year, now_tashkent().month, allowed_orgs

    now = now_tashkent()
    target_year = year or now.year
    target_month = month or now.month
    days_in_month = monthrange(target_year, target_month)[1]
    month_start = datetime(target_year, target_month, 1)
    month_end = (
        datetime(target_year + 1, 1, 1) if target_month == 12
        else datetime(target_year, target_month + 1, 1)
    )

    query = _get_employee_base_query(db, allowed_orgs, month_start, month_end)

    if organization_id is not None:
        if organization_id not in allowed_orgs:
            return [], days_in_month, target_year, target_month, allowed_orgs
        query = query.filter(Employee.organization_id == organization_id)

    if branch_id is not None:
        query = query.filter(Employee.branch_id == branch_id)
    if department_id is not None:
        query = query.filter(Employee.department_id == department_id)
    if position_id is not None:
        query = query.filter(Employee.position_id == position_id)

    if search:
        s = f"%{search.strip().lower()}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(s),
                Employee.last_name.ilike(s),
                Employee.middle_name.ilike(s),
                Employee.position.ilike(s),
            )
        )

    employees = query.order_by(Employee.last_name, Employee.first_name).all()
    if not employees:
        return [], days_in_month, target_year, target_month, allowed_orgs

    emp_ids = [emp.id for emp in employees]

    logs = (
        db.query(AttendanceLog)
        .filter(
            AttendanceLog.employee_id.in_(emp_ids),
            AttendanceLog.timestamp >= month_start,
            AttendanceLog.timestamp < month_end,
        )
        .order_by(AttendanceLog.employee_id, AttendanceLog.timestamp.asc())
        .all()
    )

    logs_by_emp_day: dict[tuple, list] = {}
    for log in logs:
        if not log.timestamp:
            continue
        day_str = log.timestamp.strftime("%Y-%m-%d")
        key = (log.employee_id, day_str)
        if key not in logs_by_emp_day:
            logs_by_emp_day[key] = [log.timestamp, log.timestamp]
        else:
            if log.timestamp < logs_by_emp_day[key][0]:
                logs_by_emp_day[key][0] = log.timestamp
            if log.timestamp > logs_by_emp_day[key][1]:
                logs_by_emp_day[key][1] = log.timestamp

    # Dinamik bayram/dam olish ma'lumotlarini yuklash
    holiday_details = load_holiday_details(
        db,
        start_date=month_start.date(),
        end_date=date(target_year, target_month, days_in_month),
        organization_ids=allowed_orgs,
    )

    today_date = now.date()
    status_by_emp = load_employee_statuses(
        db, emp_ids, month_start.date(), date(target_year, target_month, days_in_month)
    )

    tabel_data = []
    for emp in employees:
        emp_statuses = status_by_emp.get(emp.id, [])
        emp_days = []

        for day_num in range(1, days_in_month + 1):
            day_dt = date(target_year, target_month, day_num)
            day_str = day_dt.isoformat()

            is_off, off_reason = is_day_off_for_emp(day_dt, emp.organization_id, holiday_details)
            active_status = _get_active_status(emp_statuses, day_dt)
            seen_pair = logs_by_emp_day.get((emp.id, day_str))

            if active_status:
                if active_status.status_type in ("resigned", "suspended"):
                    status = active_status.status_type
                    first_seen_time = None
                    last_seen_time = None
                elif active_status.status_type in ("vacation", "sick_leave", "business_trip", "other"):
                    status = active_status.status_type
                    first_seen_time = seen_pair[0].strftime("%H:%M") if seen_pair else None
                    last_seen_time = seen_pair[1].strftime("%H:%M") if seen_pair else None
                else:
                    status = active_status.status_type
                    first_seen_time = None
                    last_seen_time = None
            elif is_off:
                status = "holiday"
                first_seen_time = seen_pair[0].strftime("%H:%M") if seen_pair else None
                last_seen_time = seen_pair[1].strftime("%H:%M") if seen_pair else None
            elif seen_pair:
                first_seen = seen_pair[0]
                last_seen = seen_pair[1]
                first_seen_time = first_seen.strftime("%H:%M")
                last_seen_time = last_seen.strftime("%H:%M")
                expected_end = get_expected_end_dt(emp, day_dt)
                if first_seen >= expected_end:
                    status = "absent"
                else:
                    late_mins = get_late_minutes(emp, day_dt, first_seen)
                    status = "late" if late_mins > 0 else "present"
            else:
                first_seen_time = None
                last_seen_time = None
                if day_dt > today_date:
                    status = "pending"
                else:
                    status = "absent"

            emp_days.append({
                "day": day_num,
                "status": status,
                "first_seen": first_seen_time,
                "last_seen": last_seen_time,
                "is_weekend": off_reason == "weekend",
                "is_holiday": off_reason == "holiday",
            })

        tabel_data.append({
            "id": emp.id,
            "uuid": emp.uuid,
            "name": f"{emp.last_name or ''} {emp.first_name or ''} {emp.middle_name or ''}".strip(),
            "position": emp.position_ref.name if emp.position_ref else (emp.position or ""),
            "days": emp_days,
        })

    return tabel_data, days_in_month, target_year, target_month, allowed_orgs


@router.get("/api/finance/tabel")
def get_tabel(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    position_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    tabel_data, days_in_month, target_year, target_month, _ = _fetch_tabel_data_internal(
        request, organization_id, branch_id, department_id, position_id, search, year, month, db
    )
    return {
        "tabel": tabel_data,
        "days_in_month": days_in_month,
        "year": target_year,
        "month": target_month,
    }


@router.get("/api/finance/tabel/export-excel")
def export_tabel_excel(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    position_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """O'zbekiston mehnat tabel standartiga muvofiq Excel eksport."""
    tabel_data, days_in_month, target_year, target_month, _ = _fetch_tabel_data_internal(
        request, organization_id, branch_id, department_id, position_id, search, year, month, db
    )

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    MONTH_NAMES = [
        "", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentyabr", "Oktyabr", "Noyabr", "Dekabr",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Tabel {target_month:02d}-{target_year}"

    header_font = Font(name="Times New Roman", size=9, bold=True, color="FFFFFF")
    cell_font = Font(name="Times New Roman", size=9)
    bold_font = Font(name="Times New Roman", size=9, bold=True)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wknd_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    late_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    vacation_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    sick_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 20

    total_cols = 3 + days_in_month + 5
    last_col_letter = ws.cell(row=1, column=total_cols).column_letter

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "O'ZBEKISTON RESPUBLIKASI MEHNAT QONUNCHILIGI TALABLARI ASOSIDA"
    ws["A1"].font = Font(name="Times New Roman", size=9, italic=True)
    ws["A1"].alignment = center

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = "XODIMLAR ISH VAQTINI HISOBGA OLISH TABELI"
    ws["A2"].font = Font(name="Times New Roman", size=14, bold=True)
    ws["A2"].alignment = center

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = f"Muddati: {MONTH_NAMES[target_month]} {target_year} yil"
    ws["A3"].font = Font(name="Times New Roman", size=11, bold=True)
    ws["A3"].alignment = center

    row = 4
    headers = ["№", "Xodim F.I.SH\n(Familiya Ism Otasining ismi)", "Lavozimi"]
    for d in range(1, days_in_month + 1):
        headers.append(str(d))
    headers += ["Keldi\n(K)", "Kechikdi\n(Kch)", "Yo'q\n(Yo)", "Dam\n(D)", "Ta'til/Kasal"]

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_idx, emp in enumerate(tabel_data, row + 1):
        days = emp["days"]
        present = sum(1 for d in days if d["status"] in ("present", "late"))
        late = sum(1 for d in days if d["status"] == "late")
        absent = sum(1 for d in days if d["status"] == "absent")
        holiday = sum(1 for d in days if d["status"] == "holiday")
        vacation_sick = sum(1 for d in days if d["status"] in ("vacation", "sick_leave", "business_trip"))

        ws.cell(row=row_idx, column=1, value=row_idx - row).alignment = center
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).font = cell_font

        name_cell = ws.cell(row=row_idx, column=2, value=emp["name"])
        name_cell.font = cell_font
        name_cell.border = border
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)

        pos_cell = ws.cell(row=row_idx, column=3, value=emp["position"])
        pos_cell.font = cell_font
        pos_cell.border = border
        pos_cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_offset, d in enumerate(days, 4):
            status = d["status"]
            cell = ws.cell(row=row_idx, column=col_offset)
            cell.border = border
            cell.alignment = center
            cell.font = cell_font

            symbol = "—"
            if status == "present":
                symbol = "K"
                cell.fill = present_fill
            elif status == "late":
                symbol = "Kch"
                cell.fill = late_fill
            elif status == "absent":
                symbol = "Yo"
                cell.fill = absent_fill
            elif status == "holiday":
                symbol = "D"
                cell.fill = wknd_fill
            elif status == "vacation":
                symbol = "Ta"
                cell.fill = vacation_fill
            elif status == "sick_leave":
                symbol = "Ka"
                cell.fill = sick_fill
            elif status == "business_trip":
                symbol = "Xs"
                cell.fill = vacation_fill
            elif status == "resigned":
                symbol = "Bo"
            elif status == "suspended":
                symbol = "Vt"
            elif status == "pending":
                symbol = ""

            cell.value = symbol

        summary = [present, late, absent, holiday, vacation_sick]
        for i, val in enumerate(summary):
            c = ws.cell(row=row_idx, column=days_in_month + 4 + i, value=val)
            c.font = bold_font
            c.alignment = center
            c.border = border

        ws.row_dimensions[row_idx].height = 16

    note_row = row + len(tabel_data) + 2
    ws.merge_cells(f"A{note_row}:{last_col_letter}{note_row}")
    ws[f"A{note_row}"] = (
        "BELGILAR: K — Keldi | Kch — Kechikib keldi | Yo — Yo'q (sababsiz) | "
        "D — Dam olish kuni | Ta — Ta'til | Ka — Kasallik | "
        "Xs — Xizmat safari | Bo — Bo'shatilgan | Vt — Vaqtincha to'xtatilgan"
    )
    ws[f"A{note_row}"].font = Font(name="Times New Roman", size=8, italic=True)

    sign_row = note_row + 2
    ws.merge_cells(f"A{sign_row}:I{sign_row}")
    ws[f"A{sign_row}"] = "Mas'ul shaxs: ________________________ (imzo)"
    ws[f"A{sign_row}"].font = cell_font

    ws.merge_cells(f"K{sign_row}:{last_col_letter}{sign_row}")
    ws[f"K{sign_row}"] = "Rahbar: ________________________ (imzo)"
    ws[f"K{sign_row}"].font = cell_font

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[ws.cell(row=4, column=d + 3).column_letter].width = 4.2
    for i in range(5):
        ws.column_dimensions[ws.cell(row=4, column=days_in_month + 4 + i).column_letter].width = 7

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"tabel_{target_year}_{target_month:02d}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/api/finance/tabel/export-pdf")
def export_tabel_pdf(
    request: Request,
    organization_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    position_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    month: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """O'zbekiston mehnat tabel standartiga muvofiq PDF eksport."""
    tabel_data, days_in_month, target_year, target_month, _ = _fetch_tabel_data_internal(
        request, organization_id, branch_id, department_id, position_id, search, year, month, db
    )

    from fpdf import FPDF

    MONTH_NAMES = [
        "", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentyabr", "Oktyabr", "Noyabr", "Dekabr",
    ]

    FONT_REGULAR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    FONT_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    has_unicode = os.path.exists(FONT_REGULAR)

    class TabelPDF(FPDF):
        def header(self):
            if has_unicode:
                self.add_font("DVS", "", FONT_REGULAR, uni=True)
                self.add_font("DVS", "B", FONT_BOLD, uni=True)
                self.set_font("DVS", "B", 12)
            else:
                self.set_font("Helvetica", "B", 12)
            self.cell(
                0, 8,
                f"XODIMLAR ISH VAQTINI HISOBGA OLISH TABELI",
                ln=True, align="C"
            )
            font = "DVS" if has_unicode else "Helvetica"
            self.set_font(font, "", 9)
            self.cell(
                0, 5,
                f"Muddati: {MONTH_NAMES[target_month]} {target_year} yil",
                ln=True, align="C"
            )
            self.ln(3)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 7)
            self.cell(0, 6, f"Bet {self.page_no()} | BioFace Tizimi", align="C")

    pdf = TabelPDF(orientation="L", unit="mm", format="A4")
    pdf.set_margins(8, 12, 8)
    pdf.add_page()

    font = "DVS" if has_unicode else "Helvetica"

    num_w = 7.0
    name_w = 42.0
    pos_w = 28.0
    sum_w = 8.0
    avail_w = 297.0 - 16.0 - num_w - name_w - pos_w - sum_w * 5
    day_w = max(3.5, avail_w / days_in_month)

    row_h = 6.0

    pdf.set_font(font, "B", 7)
    pdf.cell(num_w, row_h + 2, "№", border=1, align="C")
    pdf.cell(name_w, row_h + 2, "F.I.SH", border=1, align="C")
    pdf.cell(pos_w, row_h + 2, "Lavozimi", border=1, align="C")

    for day in range(1, days_in_month + 1):
        pdf.cell(day_w, row_h + 2, str(day), border=1, align="C")
    pdf.cell(sum_w, row_h + 2, "K", border=1, align="C")
    pdf.cell(sum_w, row_h + 2, "Kch", border=1, align="C")
    pdf.cell(sum_w, row_h + 2, "Yo", border=1, align="C")
    pdf.cell(sum_w, row_h + 2, "D", border=1, align="C")
    pdf.cell(sum_w, row_h + 2, "Ta/Ka", border=1, align="C")
    pdf.ln()

    pdf.set_font(font, "", 6.0)
    for idx, emp in enumerate(tabel_data, 1):
        days = emp["days"]
        present = sum(1 for d in days if d["status"] in ("present", "late"))
        late = sum(1 for d in days if d["status"] == "late")
        absent = sum(1 for d in days if d["status"] == "absent")
        holiday = sum(1 for d in days if d["status"] == "holiday")
        vacation_sick = sum(1 for d in days if d["status"] in ("vacation", "sick_leave", "business_trip"))

        name_str = emp["name"][:28] + ".." if len(emp["name"]) > 28 else emp["name"]
        pos_str = emp["position"][:18] + ".." if len(emp["position"]) > 18 else emp["position"]

        pdf.cell(num_w, row_h, str(idx), border=1, align="C")
        pdf.cell(name_w, row_h, name_str, border=1)
        pdf.cell(pos_w, row_h, pos_str, border=1)

        for d in days:
            status = d["status"]
            symbol = {
                "present": "K", "late": "Kch", "absent": "Yo",
                "holiday": "D", "vacation": "Ta", "sick_leave": "Ka",
                "business_trip": "Xs", "resigned": "Bo", "suspended": "Vt",
                "pending": "",
            }.get(status, "—")
            pdf.cell(day_w, row_h, symbol, border=1, align="C")

        pdf.set_font(font, "B", 7)
        pdf.cell(sum_w, row_h, str(present), border=1, align="C")
        pdf.cell(sum_w, row_h, str(late), border=1, align="C")
        pdf.cell(sum_w, row_h, str(absent), border=1, align="C")
        pdf.cell(sum_w, row_h, str(holiday), border=1, align="C")
        pdf.cell(sum_w, row_h, str(vacation_sick), border=1, align="C")
        pdf.set_font(font, "", 6.0)
        pdf.ln()

    pdf.ln(8)
    pdf.set_font(font, "B", 8)
    pdf.cell(80, 6, "Mas'ul shaxs: _____________________", ln=False)
    pdf.cell(60, 6, "", ln=False)
    pdf.cell(80, 6, "Rahbar: _____________________", ln=True)

    out = io.BytesIO()
    pdf_bytes = pdf.output(dest="S")
    out.write(pdf_bytes if isinstance(pdf_bytes, bytes) else pdf_bytes.encode("latin-1"))
    out.seek(0)

    filename = f"tabel_{target_year}_{target_month:02d}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
