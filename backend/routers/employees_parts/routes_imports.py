import threading
import uuid
from typing import Any, Optional, List
from datetime import datetime, date

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request, UploadFile, File, Form
from sqlalchemy import func
from sqlalchemy.orm import Session

from core.database import SessionLocal, get_db
from core.models import AttendanceLog, Device, Employee, EmployeeCameraLink, Organization, Branch, Department, Position, User, UserRole
from routers.cameras import (
    _resolve_online_command_target,
    _send_isup_command_or_raise,
    import_camera_users_to_db_impl,
    _save_face_bytes_to_local,
)
from routers.employees_parts.common import (
    normalize_employee_type_for_import,
    now_iso,
    resolve_allowed_org_ids,
    set_import_job,
    update_import_job,
    get_import_job,
    generate_unique_personal_id,
    normalize_personal_id,
)
from utils.face_embeddings import trigger_embedding_generation_bg

def _save_face_bytes_to_local_for_import(raw: bytes) -> Optional[str]:
    if not raw:
        return None
    import uuid
    import os
    from io import BytesIO
    from PIL import Image
    
    target_dir = "static/uploads/employees"
    try:
        with Image.open(BytesIO(raw)) as src:
            frame_count = int(getattr(src, "n_frames", 1) or 1)
            if frame_count > 1:
                return None
                
            rgb = src.convert("RGB")
            if rgb.width < 32 or rgb.height < 32:
                print(f"[IMPORT IMAGE] Image too small: {rgb.width}x{rgb.height}")
                return None
                
            max_side = 640
            if max(rgb.size) > max_side:
                rgb.thumbnail((max_side, max_side), Image.Resampling.LANCZOS)
                
            out = BytesIO()
            rgb.save(out, format="JPEG", quality=85, optimize=True)
            encoded = out.getvalue()
            
            os.makedirs(target_dir, exist_ok=True)
            filename = f"{uuid.uuid4().hex}.jpg"
            filepath = os.path.join(target_dir, filename)
            with open(filepath, "wb") as f:
                f.write(encoded)
                
            return f"/{target_dir.replace(os.sep, '/')}/{filename}"
    except Exception as e:
        print(f"[IMPORT IMAGE] Error processing image bytes: {str(e)}")
        return None

router = APIRouter()

PATRONYMIC_SUFFIXES = {"qizi", "qiz", "o'g'li", "ogli", "ugli", "ovna", "ovich", "evich", "yevich", "yevna"}


def _looks_like_surname(token: str) -> bool:
    text = str(token or "").strip().lower()
    return text.endswith(("ov", "ova", "ev", "eva", "yev", "yeva", "vich", "vna"))


def _split_import_person_name(full_name: Optional[str], fallback_personal_id: str) -> tuple[str, str, str]:
    text = str(full_name or "").strip()
    if not text:
        return "Foydalanuvchi", fallback_personal_id, ""

    parts = [p.strip() for p in text.split() if p.strip()]
    if len(parts) == 1:
        return parts[0], fallback_personal_id, ""

    has_patronymic = str(parts[-1]).lower() in PATRONYMIC_SUFFIXES
    if len(parts) >= 3 or has_patronymic:
        last_name = parts[0]
        first_name = parts[1]
        middle_name = " ".join(parts[2:])
        return first_name or "Foydalanuvchi", last_name or fallback_personal_id, middle_name

    if len(parts) == 2 and _looks_like_surname(parts[0]):
        return parts[1], parts[0], ""

    return parts[0], parts[1], ""


@router.get("/api/employees/import/sources")
def get_employees_import_sources(
    request: Request,
    organization_id: Optional[str] = Query(None),
    limit_per_camera: int = Query(1200, ge=1, le=5000),
    db: Session = Depends(get_db),
):
    allowed_org_ids = resolve_allowed_org_ids(request, db)
    if not allowed_org_ids:
        return {"ok": True, "organizations": []}

    org_query = db.query(Organization).filter(Organization.id.in_(allowed_org_ids))
    if organization_id is not None:
        org_obj = db.query(Organization).filter(Organization.uuid == str(organization_id)).first()
        if org_obj is None and str(organization_id).isdigit():
            org_obj = db.query(Organization).filter(Organization.id == int(organization_id)).first()
        if org_obj is not None:
            org_query = org_query.filter(Organization.id == org_obj.id)
        else:
            org_query = org_query.filter(Organization.id == -1)
    orgs = org_query.order_by(Organization.name.asc()).all()

    payload_orgs: list[dict[str, Any]] = []
    for org in orgs:
        cameras = db.query(Device).filter(Device.organization_id == org.id).order_by(Device.name.asc()).all()
        cam_rows: list[dict[str, Any]] = []
        for cam in cameras:
            count = 0
            unsupported = False
            error_msg = None
            try:
                target_id, _, _ = _resolve_online_command_target(cam)
                count_resp = _send_isup_command_or_raise(target_id, "get_face_count", {}, timeout=12.0)
                count = int(
                    count_resp.get("face_count")
                    or count_resp.get("bind_face_user_count")
                    or count_resp.get("fd_record_total")
                    or 0
                )
                if count <= 0:
                    face_resp = _send_isup_command_or_raise(
                        target_id,
                        "get_face_records",
                        {"all": True, "limit": limit_per_camera},
                        timeout=20.0,
                    )
                    rows = face_resp.get("records", []) if isinstance(face_resp, dict) else []
                    count = len(rows) if isinstance(rows, list) else 0
                
                if cam.used_faces != count:
                    cam.used_faces = count
                    db.commit()
            except HTTPException as exc:
                txt = str(exc.detail or "")
                unsupported = "notsupport" in txt.lower()
                error_msg = txt

            cam_rows.append(
                {
                    "id": int(cam.id),
                    "name": str(cam.name or f"Kamera #{cam.id}"),
                    "organization_id": int(org.id),
                    "organization_name": str(org.name or ""),
                    "user_count": int(count),
                    "unsupported": bool(unsupported),
                    "error": error_msg,
                }
            )

        payload_orgs.append(
            {
                "id": int(org.id),
                "name": str(org.name or ""),
                "camera_count": len(cam_rows),
                "cameras": cam_rows,
            }
        )

    return {"ok": True, "organizations": payload_orgs}


def _run_employees_import_job(*, job_id: str, camera_ids: list[int], employee_type: Optional[str]):
    db = SessionLocal()
    try:
        total = len(camera_ids)
        update_import_job(job_id, status="running", total_cameras=total, started_at=now_iso())

        summary = {
            "created": 0,
            "updated": 0,
            "existing": 0,
            "linked_to_camera": 0,
            "already_linked": 0,
            "skipped": 0,
            "imported_users_total": 0,
        }
        per_camera: list[dict[str, Any]] = []
        processed_items = 0
        total_items = 0

        for idx, cam_id in enumerate(camera_ids, start=1):
            update_import_job(
                job_id,
                current_camera_id=int(cam_id),
                current_camera_name=f"Kamera #{cam_id}",
                processed_cameras=idx - 1,
                progress_percent=int(((idx - 1) / max(1, total)) * 100),
                processed_items=int(processed_items),
                total_items=int(total_items),
                current_personal_id="",
                progress_note="",
                heartbeat_at=now_iso(),
            )
            try:
                def _progress(done: int, total_for_camera: int, current_personal_id: Optional[str], partial_summary: dict[str, Any]) -> None:
                    known_total = processed_items + max(0, int(total_for_camera or 0))
                    current_camera_name = str(partial_summary.get("camera_name") or f"Kamera #{cam_id}")
                    progress_note = str(partial_summary.get("progress_note") or "")
                    live_summary = {
                        "created": int(summary.get("created") or 0) + int(partial_summary.get("created") or 0),
                        "updated": int(summary.get("updated") or 0) + int(partial_summary.get("updated") or 0),
                        "existing": int(summary.get("existing") or 0) + int(partial_summary.get("existing") or 0),
                        "linked_to_camera": int(summary.get("linked_to_camera") or 0) + int(partial_summary.get("linked_to_camera") or 0),
                        "already_linked": int(summary.get("already_linked") or 0) + int(partial_summary.get("already_linked") or 0),
                        "skipped": int(summary.get("skipped") or 0) + int(partial_summary.get("skipped") or 0),
                        "imported_users_total": int(summary.get("imported_users_total") or 0) + int(partial_summary.get("imported_users_total") or 0),
                    }
                    live_per_camera = list(per_camera)
                    live_per_camera.append(
                        {
                            "camera_id": int(partial_summary.get("camera_id") or cam_id),
                            "camera_name": current_camera_name,
                            "ok": True,
                            "created": int(partial_summary.get("created") or 0),
                            "updated": int(partial_summary.get("updated") or 0),
                            "existing": int(partial_summary.get("existing") or 0),
                            "linked_to_camera": int(partial_summary.get("linked_to_camera") or 0),
                            "already_linked": int(partial_summary.get("already_linked") or 0),
                            "skipped": int(partial_summary.get("skipped") or 0),
                            "imported_users_total": int(partial_summary.get("imported_users_total") or 0),
                            "message": progress_note or "Import davom etmoqda",
                        }
                    )
                    update_import_job(
                        job_id,
                        current_camera_id=int(cam_id),
                        current_camera_name=current_camera_name,
                        processed_cameras=idx - 1,
                        processed_items=int(processed_items + max(0, int(done or 0))),
                        total_items=int(known_total),
                        current_personal_id=str(current_personal_id or ""),
                        progress_note=progress_note,
                        progress_percent=int(((processed_items + max(0, int(done or 0))) / max(1, known_total)) * 100) if known_total > 0 else 0,
                        summary=live_summary,
                        per_camera=live_per_camera,
                        heartbeat_at=now_iso(),
                    )

                result = import_camera_users_to_db_impl(
                    cam_id=int(cam_id),
                    limit=5000,
                    allow_camera_http_download=False,
                    face_import_mode="off",
                    employee_type=employee_type,
                    only_with_face=True,
                    prefer_face_records_only=True,
                    progress_cb=_progress,
                    db=db,
                )
                imported_total = int(result.get("imported_users_total") or 0)
                processed_items += imported_total
                total_items = max(total_items, processed_items)
                cam_info = {
                    "camera_id": int(cam_id),
                    "camera_name": str(result.get("camera_name") or f"Kamera #{cam_id}"),
                    "ok": bool(result.get("ok", True)),
                    "created": int(result.get("created") or 0),
                    "updated": int(result.get("updated") or 0),
                    "existing": int(result.get("existing") or 0),
                    "linked_to_camera": int(result.get("linked_to_camera") or 0),
                    "already_linked": int(result.get("already_linked") or 0),
                    "skipped": int(result.get("skipped") or 0),
                    "imported_users_total": int(result.get("imported_users_total") or 0),
                    "message": str(result.get("message") or ""),
                }
                per_camera.append(cam_info)
                for key in summary:
                    summary[key] += int(cam_info.get(key) or 0)
            except Exception as exc:
                per_camera.append(
                    {
                        "camera_id": int(cam_id),
                        "camera_name": f"Kamera #{cam_id}",
                        "ok": False,
                        "created": 0,
                        "updated": 0,
                        "existing": 0,
                        "linked_to_camera": 0,
                        "already_linked": 0,
                        "skipped": 0,
                        "imported_users_total": 0,
                        "message": str(exc),
                    }
                )

            update_import_job(
                job_id,
                processed_cameras=idx,
                current_camera_name=str(per_camera[-1].get("camera_name") or f"Kamera #{cam_id}") if per_camera else f"Kamera #{cam_id}",
                processed_items=int(processed_items),
                total_items=int(total_items),
                current_personal_id="",
                progress_note="",
                progress_percent=int((processed_items / max(1, total_items)) * 100) if total_items > 0 else int((idx / max(1, total)) * 100),
                summary=summary,
                per_camera=per_camera,
                heartbeat_at=now_iso(),
            )

        update_import_job(
            job_id,
            status="done",
            progress_percent=100,
            finished_at=now_iso(),
            summary=summary,
            per_camera=per_camera,
            processed_items=int(processed_items),
            total_items=int(total_items),
            current_camera_name=str(per_camera[-1].get("camera_name") or "") if per_camera else "",
            current_personal_id="",
            progress_note="",
        )
    except Exception as exc:
        update_import_job(job_id, status="error", error=str(exc), finished_at=now_iso())
    finally:
        db.close()


@router.post("/api/employees/import/start")
def start_employees_import_job(
    request: Request,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    camera_ids_raw = payload.get("camera_ids") if isinstance(payload, dict) else []
    employee_type = normalize_employee_type_for_import(payload.get("employee_type") if isinstance(payload, dict) else None)
    if not isinstance(camera_ids_raw, list):
        raise HTTPException(status_code=422, detail="camera_ids ro'yxat bo'lishi kerak")
    camera_ids = []
    for cam_id in camera_ids_raw:
        try:
            val = int(cam_id)
        except Exception:
            continue
        if val > 0 and val not in camera_ids:
            camera_ids.append(val)
    if not camera_ids:
        raise HTTPException(status_code=422, detail="Kamida bitta kamera tanlang")

    allowed_org_ids = resolve_allowed_org_ids(request, db)
    cams = db.query(Device.id, Device.organization_id).filter(Device.id.in_(camera_ids)).all()
    cam_map = {int(row.id): (int(row.organization_id) if row.organization_id is not None else None) for row in cams}
    valid_camera_ids = [cid for cid in camera_ids if cid in cam_map and cam_map[cid] in allowed_org_ids]
    if not valid_camera_ids:
        raise HTTPException(status_code=403, detail="Tanlangan kameralar uchun ruxsat yo'q")

    job_id = uuid.uuid4().hex
    set_import_job(
        job_id,
        {
            "job_id": job_id,
            "kind": "camera_import",
            "status": "queued",
            "created_at": now_iso(),
            "processed_cameras": 0,
            "total_cameras": len(valid_camera_ids),
            "progress_percent": 0,
            "processed_items": 0,
            "total_items": 0,
            "current_camera_id": None,
            "current_camera_name": "",
            "current_personal_id": "",
            "progress_note": "",
            "summary": {
                "created": 0,
                "updated": 0,
                "existing": 0,
                "linked_to_camera": 0,
                "already_linked": 0,
                "skipped": 0,
                "imported_users_total": 0,
            },
            "per_camera": [],
            "employee_type": employee_type,
        },
    )

    thread = threading.Thread(
        target=_run_employees_import_job,
        kwargs={"job_id": job_id, "camera_ids": valid_camera_ids, "employee_type": employee_type},
        daemon=True,
        name=f"employees-import-{job_id[:8]}",
    )
    thread.start()
    return {"ok": True, "job_id": job_id, "message": "Import jarayoni boshlandi"}


@router.get("/api/employees/import/status")
def get_employees_import_status(job_id: str):
    state = get_import_job(str(job_id or "").strip())
    if not state:
        raise HTTPException(status_code=404, detail="Import job topilmadi")
    return {"ok": True, **state}


def _import_from_attendance_core(
    db: Session,
    *,
    allowed_org_ids: list[int],
    org_id_int: Optional[int],
    cam_id_int: Optional[int],
    employee_type: Optional[str],
    limit: int,
    progress_cb: Optional[Any] = None,
) -> dict[str, Any]:
    logs_query = (
        db.query(AttendanceLog)
        .outerjoin(AttendanceLog.device)
        .filter(
            AttendanceLog.person_id.isnot(None),
            func.trim(AttendanceLog.person_id) != "",
            Device.organization_id.in_(allowed_org_ids),
        )
    )
    if org_id_int is not None:
        logs_query = logs_query.filter(Device.organization_id == org_id_int)
    if cam_id_int is not None:
        logs_query = logs_query.filter(AttendanceLog.device_id == cam_id_int)

    logs = logs_query.order_by(AttendanceLog.timestamp.desc(), AttendanceLog.id.desc()).limit(limit).all()
    grouped: dict[str, dict[str, Any]] = {}
    for log in logs:
        pid = str(log.person_id or "").strip()
        if not pid:
            continue
        entry = grouped.setdefault(
            pid,
            {
                "name": str(log.person_name or "").strip(),
                "camera_ids": set(),
                "organization_id": log.device.organization_id if log.device else None,
                "log_ids": [],
            },
        )
        if not entry["name"] and log.person_name:
            entry["name"] = str(log.person_name)
        if entry["organization_id"] is None and log.device and log.device.organization_id is not None:
            entry["organization_id"] = log.device.organization_id
        if log.device_id is not None:
            entry["camera_ids"].add(int(log.device_id))
        entry["log_ids"].append(int(log.id))

    created = 0
    updated = 0
    linked_to_camera = 0
    logs_bound = 0
    total_users = len(grouped)

    for idx, (personal_id, info) in enumerate(grouped.items(), start=1):
        first_name, last_name, middle_name = _split_import_person_name(info.get("name"), personal_id)
        emp = db.query(Employee).filter(Employee.personal_id == personal_id).first()
        if emp is None:
            emp = Employee(
                first_name=first_name,
                last_name=last_name,
                middle_name=middle_name,
                personal_id=personal_id,
                employee_type=employee_type,
                has_access=True,
                organization_id=info.get("organization_id"),
            )
            db.add(emp)
            db.flush()
            created += 1
        else:
            changed = False
            if first_name and emp.first_name != first_name:
                emp.first_name = first_name
                changed = True
            if last_name and emp.last_name != last_name:
                emp.last_name = last_name
                changed = True
            if middle_name and emp.middle_name != middle_name:
                emp.middle_name = middle_name
                changed = True
            if employee_type is not None and emp.employee_type != employee_type:
                emp.employee_type = employee_type
                changed = True
            if emp.organization_id is None and info.get("organization_id") is not None:
                emp.organization_id = info.get("organization_id")
                changed = True
            if changed:
                updated += 1

        for cam_id in sorted(list(info.get("camera_ids") or [])):
            exists = (
                db.query(EmployeeCameraLink.id)
                .filter(EmployeeCameraLink.employee_id == int(emp.id), EmployeeCameraLink.camera_id == int(cam_id))
                .first()
            )
            if not exists:
                db.add(EmployeeCameraLink(employee_id=int(emp.id), camera_id=int(cam_id)))
                linked_to_camera += 1

        if info.get("log_ids"):
            updated_rows = (
                db.query(AttendanceLog)
                .filter(AttendanceLog.id.in_(info["log_ids"]), AttendanceLog.employee_id.is_(None))
                .update({AttendanceLog.employee_id: int(emp.id)}, synchronize_session=False)
            )
            logs_bound += int(updated_rows or 0)

        if callable(progress_cb):
            progress_cb(idx, total_users, personal_id)

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "linked_to_camera": linked_to_camera,
        "logs_bound": logs_bound,
        "processed_users": total_users,
    }


def _run_attendance_import_job(
    *,
    job_id: str,
    allowed_org_ids: list[int],
    org_id_int: Optional[int],
    cam_id_int: Optional[int],
    employee_type: Optional[str],
    limit: int,
):
    db = SessionLocal()
    try:
        update_import_job(job_id, status="running", started_at=now_iso(), heartbeat_at=now_iso())

        def _progress(done: int, total: int, current_personal_id: str) -> None:
            update_import_job(
                job_id,
                processed_items=int(done),
                total_items=int(total),
                current_personal_id=str(current_personal_id or ""),
                progress_percent=int((int(done) / max(1, int(total))) * 100),
                heartbeat_at=now_iso(),
            )

        result = _import_from_attendance_core(
            db,
            allowed_org_ids=allowed_org_ids,
            org_id_int=org_id_int,
            cam_id_int=cam_id_int,
            employee_type=employee_type,
            limit=limit,
            progress_cb=_progress,
        )
        update_import_job(
            job_id,
            status="done",
            progress_percent=100,
            finished_at=now_iso(),
            summary=result,
            processed_items=int(result.get("processed_users") or 0),
            total_items=int(result.get("processed_users") or 0),
        )
    except Exception as exc:
        update_import_job(job_id, status="error", error=str(exc), finished_at=now_iso())
    finally:
        db.close()


@router.post("/api/employees/import/from-attendance/start")
def start_import_from_attendance_job(
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    data = payload if isinstance(payload, dict) else {}
    organization_id = data.get("organization_id")
    camera_id = data.get("camera_id")
    employee_type = normalize_employee_type_for_import(data.get("employee_type"))
    limit = max(1, min(int(data.get("limit") or 5000), 50000))

    try:
        org_id_int = int(organization_id) if organization_id is not None else None
    except Exception as exc:
        raise HTTPException(status_code=422, detail="organization_id noto'g'ri") from exc
    try:
        cam_id_int = int(camera_id) if camera_id is not None else None
    except Exception as exc:
        raise HTTPException(status_code=422, detail="camera_id noto'g'ri") from exc

    allowed_org_ids = resolve_allowed_org_ids(request, db)
    if not allowed_org_ids:
        raise HTTPException(status_code=403, detail="Import uchun ruxsatli tashkilot topilmadi")
    if org_id_int is not None and org_id_int not in allowed_org_ids:
        raise HTTPException(status_code=403, detail="Tashkilot uchun ruxsat yo'q")
    if cam_id_int is not None:
        cam = db.query(Device.id, Device.organization_id).filter(Device.id == cam_id_int).first()
        if not cam:
            raise HTTPException(status_code=404, detail="Kamera topilmadi")
        if cam.organization_id not in allowed_org_ids:
            raise HTTPException(status_code=403, detail="Kamera uchun ruxsat yo'q")

    job_id = uuid.uuid4().hex
    set_import_job(
        job_id,
        {
            "job_id": job_id,
            "kind": "attendance_import",
            "status": "queued",
            "created_at": now_iso(),
            "progress_percent": 0,
            "processed_items": 0,
            "total_items": 0,
            "current_personal_id": "",
            "summary": {
                "created": 0,
                "updated": 0,
                "linked_to_camera": 0,
                "logs_bound": 0,
                "processed_users": 0,
            },
        },
    )

    thread = threading.Thread(
        target=_run_attendance_import_job,
        kwargs={
            "job_id": job_id,
            "allowed_org_ids": allowed_org_ids,
            "org_id_int": org_id_int,
            "cam_id_int": cam_id_int,
            "employee_type": employee_type,
            "limit": limit,
        },
        daemon=True,
        name=f"attendance-import-{job_id[:8]}",
    )
    thread.start()

    return {"ok": True, "job_id": job_id, "message": "Davomatdan import jarayoni boshlandi"}


@router.get("/api/employees/import/from-attendance/sources")
def get_attendance_import_sources(request: Request, db: Session = Depends(get_db)):
    allowed_org_ids = resolve_allowed_org_ids(request, db)
    if not allowed_org_ids:
        return {"ok": True, "organizations": []}

    grouped_rows = (
        db.query(
            Device.organization_id.label("org_id"),
            Device.id.label("camera_id"),
            Device.name.label("camera_name"),
            func.count(func.distinct(func.trim(AttendanceLog.person_id))).label("users_count"),
        )
        .join(Device, Device.id == AttendanceLog.device_id)
        .filter(
            Device.organization_id.in_(allowed_org_ids),
            AttendanceLog.person_id.isnot(None),
            func.trim(AttendanceLog.person_id) != "",
        )
        .group_by(Device.organization_id, Device.id, Device.name)
        .order_by(Device.organization_id.asc(), Device.name.asc())
        .all()
    )

    org_names = {
        int(row.id): str(row.name or "")
        for row in db.query(Organization.id, Organization.name).filter(Organization.id.in_(allowed_org_ids)).all()
    }

    org_map: dict[int, dict[str, Any]] = {}
    for row in grouped_rows:
        org_id = int(row.org_id) if row.org_id is not None else 0
        if org_id <= 0:
            continue
        org_payload = org_map.setdefault(
            org_id,
            {
                "id": org_id,
                "name": org_names.get(org_id, f"Tashkilot #{org_id}"),
                "camera_count": 0,
                "users_count": 0,
                "cameras": [],
            },
        )
        users_count = int(row.users_count or 0)
        org_payload["cameras"].append(
            {"id": int(row.camera_id), "name": str(row.camera_name or f"Kamera #{row.camera_id}"), "users_count": users_count}
        )
        org_payload["camera_count"] += 1
        org_payload["users_count"] += users_count

    organizations = list(org_map.values())
    organizations.sort(key=lambda item: str(item.get("name") or ""))
    return {"ok": True, "organizations": organizations}


@router.post("/api/employees/import/from-attendance")
def import_employees_from_attendance(
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    data = payload if isinstance(payload, dict) else {}
    organization_id = data.get("organization_id")
    camera_id = data.get("camera_id")
    employee_type = normalize_employee_type_for_import(data.get("employee_type"))
    limit = max(1, min(int(data.get("limit") or 5000), 50000))

    try:
        org_id_int = int(organization_id) if organization_id is not None else None
    except Exception as exc:
        raise HTTPException(status_code=422, detail="organization_id noto'g'ri") from exc
    try:
        cam_id_int = int(camera_id) if camera_id is not None else None
    except Exception as exc:
        raise HTTPException(status_code=422, detail="camera_id noto'g'ri") from exc

    allowed_org_ids = resolve_allowed_org_ids(request, db)
    if not allowed_org_ids:
        return {
            "ok": True,
            "created": 0,
            "updated": 0,
            "linked_to_camera": 0,
            "logs_bound": 0,
            "processed_users": 0,
            "message": "Import uchun ruxsatli tashkilot topilmadi",
        }

    if org_id_int is not None and org_id_int not in allowed_org_ids:
        raise HTTPException(status_code=403, detail="Tashkilot uchun ruxsat yo'q")

    if cam_id_int is not None:
        cam = db.query(Device.id, Device.organization_id).filter(Device.id == cam_id_int).first()
        if not cam:
            raise HTTPException(status_code=404, detail="Kamera topilmadi")
        if cam.organization_id not in allowed_org_ids:
            raise HTTPException(status_code=403, detail="Kamera uchun ruxsat yo'q")

    result = _import_from_attendance_core(
        db,
        allowed_org_ids=allowed_org_ids,
        org_id_int=org_id_int,
        cam_id_int=cam_id_int,
        employee_type=employee_type,
        limit=limit,
    )
    return {
        "ok": True,
        **result,
        "message": f"Davomatdan import yakunlandi: {result['created']} yangi, {result['updated']} yangilandi, {result['linked_to_camera']} bog'lanish yaratildi.",
    }


@router.post("/api/employees/bulk-import")
def bulk_import_employees(
    request: Request,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    auth_user = request.session.get("auth_user") or {}
    org_id = payload.get("organization_id")
    if not org_id:
        raise HTTPException(status_code=422, detail="Tashkilot tanlanishi shart")
    
    from routers.employees_parts.common import resolve_effective_org_id
    resolved_org_id = resolve_effective_org_id(request, db, int(org_id))
    if not resolved_org_id:
        raise HTTPException(status_code=403, detail="Tashkilotga kirish ruxsati yo'q")

    branch_id = payload.get("branch_id")
    if not branch_id:
        branches = db.query(Branch).filter(Branch.organization_id == resolved_org_id).order_by(Branch.id).all()
        if not branches:
            org_obj = db.query(Organization).filter(Organization.id == resolved_org_id).first()
            default_branch = Branch(
                organization_id=resolved_org_id,
                name="Asosiy filial",
                address=org_obj.address if org_obj else None,
                latitude=org_obj.latitude if org_obj else None,
                longitude=org_obj.longitude if org_obj else None,
                radius=(org_obj.radius or 100) if org_obj else 100,
            )
            db.add(default_branch)
            db.commit()
            db.refresh(default_branch)
            branch_id = default_branch.id
        else:
            branch_id = branches[0].id
    else:
        branch_id = int(branch_id)

    items = payload.get("items") or []
    if not items:
        return {"ok": True, "imported_count": 0, "message": "Import qilinadigan xodimlar mavjud emas"}

    dept_cache = {}
    pos_cache = {}
    
    imported_count = 0
    errors = []
    
    for idx, item in enumerate(items):
        try:
            first_name = str(item.get("first_name") or item.get("Ism") or "").strip()
            last_name = str(item.get("last_name") or item.get("Familiya") or "").strip()
            if not first_name or not last_name:
                errors.append(f"Qator #{idx + 1}: Ism va Familiya bo'lishi shart")
                continue
                
            personal_id = str(item.get("personal_id") or item.get("Shaxsiy ID (Personal ID)") or "").strip()
            if personal_id:
                dup = db.query(Employee).filter(
                    Employee.organization_id == resolved_org_id,
                    Employee.personal_id == personal_id
                ).first()
                if dup:
                    errors.append(f"Qator #{idx + 1}: Personal ID '{personal_id}' allaqachon mavjud")
                    continue
            
            # Department
            dept_name = str(item.get("department") or item.get("Bo'lim") or item.get("Sinf / Guruh") or "").strip()
            dept_id = None
            if dept_name:
                if dept_name not in dept_cache:
                    dept = db.query(Department).filter(
                        Department.organization_id == resolved_org_id,
                        func.lower(Department.name) == dept_name.lower()
                    ).first()
                    if not dept:
                        dept = Department(organization_id=resolved_org_id, name=dept_name)
                        db.add(dept)
                        db.commit()
                        db.refresh(dept)
                    dept_cache[dept_name] = dept.id
                dept_id = dept_cache[dept_name]
                
            # Position
            pos_name = str(item.get("position") or item.get("Lavozim") or "").strip()
            pos_id = None
            if pos_name:
                if pos_name not in pos_cache:
                    pos = db.query(Position).filter(
                        Position.organization_id == resolved_org_id,
                        func.lower(Position.name) == pos_name.lower()
                    ).first()
                    if not pos:
                        pos = Position(organization_id=resolved_org_id, name=pos_name)
                        db.add(pos)
                        db.commit()
                        db.refresh(pos)
                    pos_cache[pos_name] = pos.id
                pos_id = pos_cache[pos_name]
            
            emp_type = str(item.get("employee_type") or item.get("Xodim turi (oqituvchi/hodim)") or item.get("O'quvchi turi (oquvchi/talaba)") or payload.get("employee_type") or "staff").strip().lower()
            if emp_type in ("oquvchi", "talaba", "students", "student"):
                emp_type = "student"
            else:
                emp_type = "staff"

            birth_date = None
            raw_bdate = str(item.get("birth_date") or item.get("Tug'ilgan sana (YYYY-MM-DD)") or "").strip()
            if raw_bdate:
                try:
                    birth_date = datetime.strptime(raw_bdate, "%Y-%m-%d")
                except:
                    pass

            emp = Employee(
                organization_id=resolved_org_id,
                branch_id=branch_id,
                first_name=first_name,
                last_name=last_name,
                middle_name=str(item.get("middle_name") or item.get("Otasining ismi") or "").strip() or None,
                personal_id=personal_id or None,
                phone=str(item.get("phone") or item.get("Telefon raqami") or "").strip() or None,
                parent_phone=str(item.get("parent_phone") or item.get("Ota-onasining telefon raqami") or "").strip() or None,
                employee_type=emp_type,
                department_id=dept_id,
                position_id=pos_id,
                region=str(item.get("region") or item.get("Viloyat") or "").strip() or None,
                district=str(item.get("district") or item.get("Tuman") or "").strip() or None,
                address=str(item.get("address") or item.get("Manzil") or "").strip() or None,
                birth_date=birth_date,
                gender=str(item.get("gender") or item.get("Jinsi (male/female)") or "").strip().lower() or None,
                has_access=True,
            )
            db.add(emp)
            imported_count += 1
        except Exception as e:
            errors.append(f"Qator #{idx + 1}: Kutilmagan xato: {str(e)}")

    db.commit()
    return {"ok": True, "imported_count": imported_count, "errors": errors}


@router.post("/api/employees/bulk-import-zip")
async def bulk_import_employees_zip(
    request: Request,
    file: UploadFile = File(...),
    organization_id: int = Form(...),
    employee_type: str = Form("staff"),
    branch_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
):
    import zipfile
    import io
    import openpyxl

    from routers.employees_parts.common import resolve_effective_org_id
    resolved_org_id = resolve_effective_org_id(request, db, int(organization_id))
    if not resolved_org_id:
        raise HTTPException(status_code=403, detail="Tashkilotga kirish ruxsati yo'q")

    if not branch_id:
        branches = db.query(Branch).filter(Branch.organization_id == resolved_org_id).order_by(Branch.id).all()
        if not branches:
            org_obj = db.query(Organization).filter(Organization.id == resolved_org_id).first()
            default_branch = Branch(
                organization_id=resolved_org_id,
                name="Asosiy filial",
                address=org_obj.address if org_obj else None,
                latitude=org_obj.latitude if org_obj else None,
                longitude=org_obj.longitude if org_obj else None,
                radius=(org_obj.radius or 100) if org_obj else 100,
            )
            db.add(default_branch)
            db.commit()
            db.refresh(default_branch)
            branch_id = default_branch.id
        else:
            branch_id = branches[0].id
    else:
        branch_id = int(branch_id)

    zip_bytes = await file.read()
    try:
        z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fayl to'g'ri ZIP arxiv emas: {str(e)}")

    namelist = z.namelist()

    excel_file_path = None
    for name in namelist:
        if "__MACOSX" in name or name.split('/')[-1].startswith('.'):
            continue
        if name.lower().endswith(('.xlsx', '.xls')):
            excel_file_path = name
            break

    if not excel_file_path:
        raise HTTPException(status_code=400, detail="ZIP arxivi ichida Excel (.xlsx, .xls) fayli topilmadi")

    try:
        excel_data = z.read(excel_file_path)
        wb = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True)
        sheet = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel faylini o'qishda xatolik: {str(e)}")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {"ok": True, "imported_count": 0, "message": "Excel faylida ma'lumot topilmadi"}

    headers = [str(cell).strip() if cell is not None else f"col_{idx}" for idx, cell in enumerate(rows[0])]
    items = []
    for r in rows[1:]:
        if all(cell is None for cell in r):
            continue
        row_dict = {}
        for idx, val in enumerate(r):
            if idx < len(headers):
                row_dict[headers[idx]] = val
        items.append(row_dict)

    dept_cache = {}
    pos_cache = {}
    imported_count = 0
    errors = []

    for idx, item in enumerate(items):
        try:
            first_name = str(item.get("first_name") or item.get("Ism") or "").strip()
            last_name = str(item.get("last_name") or item.get("Familiya") or "").strip()
            if not first_name or not last_name:
                errors.append(f"Qator #{idx + 2}: Ism va Familiya bo'lishi shart")
                continue

            personal_id = str(item.get("personal_id") or item.get("Shaxsiy ID (Personal ID)") or "").strip()
            normalized_personal_id = normalize_personal_id(personal_id) if personal_id else None
            if not normalized_personal_id:
                normalized_personal_id = generate_unique_personal_id(db)
            else:
                dup = db.query(Employee).filter(
                    Employee.organization_id == resolved_org_id,
                    Employee.personal_id == normalized_personal_id
                ).first()
                if dup:
                    errors.append(f"Qator #{idx + 2}: Personal ID '{normalized_personal_id}' allaqachon mavjud")
                    continue

            # Department
            dept_name = str(item.get("department") or item.get("Bo'lim") or item.get("Sinf / Guruh") or "").strip()
            dept_id = None
            if dept_name:
                if dept_name not in dept_cache:
                    dept = db.query(Department).filter(
                        Department.organization_id == resolved_org_id,
                        func.lower(Department.name) == dept_name.lower()
                    ).first()
                    if not dept:
                        dept = Department(organization_id=resolved_org_id, name=dept_name)
                        db.add(dept)
                        db.commit()
                        db.refresh(dept)
                    dept_cache[dept_name] = dept.id
                dept_id = dept_cache[dept_name]

            # Position
            pos_name = str(item.get("position") or item.get("Lavozim") or "").strip()
            pos_id = None
            if pos_name:
                if pos_name not in pos_cache:
                    pos = db.query(Position).filter(
                        Position.organization_id == resolved_org_id,
                        func.lower(Position.name) == pos_name.lower()
                    ).first()
                    if not pos:
                        pos = Position(organization_id=resolved_org_id, name=pos_name, department_id=dept_id)
                        db.add(pos)
                        db.commit()
                        db.refresh(pos)
                    pos_cache[pos_name] = pos.id
                pos_id = pos_cache[pos_name]

            emp_type = str(item.get("employee_type") or item.get("Xodim turi (oqituvchi/hodim)") or item.get("O'quvchi turi (oquvchi/talaba)") or employee_type or "staff").strip().lower()
            if emp_type in ("oquvchi", "talaba", "students", "student"):
                emp_type = "student"
            else:
                emp_type = "staff"

            birth_date = None
            raw_bdate_cell = item.get("birth_date") or item.get("Tug'ilgan sana (YYYY-MM-DD)")
            if isinstance(raw_bdate_cell, (datetime, date)):
                birth_date = raw_bdate_cell
            elif raw_bdate_cell:
                raw_bdate = str(raw_bdate_cell).strip()
                if raw_bdate:
                    if raw_bdate.replace('.', '', 1).isdigit():
                        try:
                            from datetime import date as d_class, timedelta
                            serial = float(raw_bdate)
                            birth_date = d_class(1899, 12, 30) + timedelta(days=serial)
                        except:
                            pass
                    else:
                        try:
                            birth_date = datetime.strptime(raw_bdate, "%Y-%m-%d")
                        except:
                            try:
                                birth_date = datetime.strptime(raw_bdate, "%d.%m.%Y")
                            except:
                                pass

            # Try to resolve photo from ZIP
            avatar_url = None
            img_filename = str(item.get("Rasm (URL yoki fayl nomi)") or item.get("Rasm") or item.get("image_url") or "").strip()
            if img_filename:
                base_name = img_filename.lower().split('/')[-1].split('\\')[-1]
                for name in namelist:
                    if "__MACOSX" in name or name.split('/')[-1].startswith('.'):
                        continue
                    zip_base_name = name.lower().split('/')[-1].split('\\')[-1]
                    if zip_base_name == base_name:
                        try:
                            image_bytes = z.read(name)
                            if image_bytes:
                                avatar_url = _save_face_bytes_to_local_for_import(image_bytes)
                        except Exception as img_err:
                            errors.append(f"Qator #{idx + 2}: Rasmni o'qishda xato ({img_filename}): {str(img_err)}")
                        break

            emp = Employee(
                organization_id=resolved_org_id,
                branch_id=branch_id,
                first_name=first_name,
                last_name=last_name,
                middle_name=str(item.get("middle_name") or item.get("Otasining ismi") or "").strip() or None,
                personal_id=normalized_personal_id,
                phone=str(item.get("phone") or item.get("Telefon raqami") or "").strip() or None,
                parent_phone=str(item.get("parent_phone") or item.get("Ota-onasining telefon raqami") or "").strip() or None,
                employee_type=emp_type,
                department_id=dept_id,
                position_id=pos_id,
                region=str(item.get("region") or item.get("Viloyat") or "").strip() or None,
                district=str(item.get("district") or item.get("Tuman") or "").strip() or None,
                address=str(item.get("address") or item.get("Manzil") or "").strip() or None,
                birth_date=birth_date,
                gender=str(item.get("gender") or item.get("Jinsi (male/female)") or "").strip().lower() or None,
                image_url=avatar_url,
                has_access=True,
            )
            db.add(emp)
            db.commit()
            db.refresh(emp)

            # Auto-create user account: username = personal_id, password = 'bioface'
            if emp.personal_id:
                try:
                    import bcrypt
                    existing_user = db.query(User).filter(User.name == str(emp.personal_id)).first()
                    if not existing_user:
                        hashed_pw = bcrypt.hashpw(b'bioface', bcrypt.gensalt()).decode('utf-8')
                        auto_user = User(
                            name=str(emp.personal_id),
                            email=f"{emp.personal_id}@bioface.local",
                            hashed_password=hashed_pw,
                            role=UserRole.tashkilot_admin,
                            status="active",
                            is_staff=False,
                            organization_id=resolved_org_id,
                            branch_id=branch_id,
                        )
                        db.add(auto_user)
                        db.commit()
                except Exception as _ue:
                    print(f"[AUTO USER] Failed to create user for employee {emp.personal_id}: {_ue}")

            if emp.image_url:
                trigger_embedding_generation_bg(employee_id=int(emp.id))

            imported_count += 1
        except Exception as e:
            errors.append(f"Qator #{idx + 2}: Kutilmagan xato: {str(e)}")

    return {"ok": True, "imported_count": imported_count, "errors": errors}


@router.post("/api/employees/clear")
def clear_employees(
    request: Request,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
):
    organization_id = payload.get("organization_id")
    if not organization_id:
        raise HTTPException(status_code=422, detail="Tashkilot tanlanishi shart")

    from routers.employees_parts.common import resolve_effective_org_id
    resolved_org_id = resolve_effective_org_id(request, db, int(organization_id))
    if not resolved_org_id:
        raise HTTPException(status_code=403, detail="Tashkilotga kirish ruxsati yo'q")

    employee_type = payload.get("employee_type")
    
    query = db.query(Employee).filter(Employee.organization_id == resolved_org_id)
    if employee_type:
        emp_type = str(employee_type).strip().lower()
        if emp_type in ("oquvchi", "talaba", "students", "student"):
            query = query.filter(Employee.employee_type == "student")
        else:
            query = query.filter(Employee.employee_type == "staff")
            
    emp_ids = [e.id for e in query.all()]
    if emp_ids:
        db.query(EmployeeCameraLink).filter(EmployeeCameraLink.employee_id.in_(emp_ids)).delete(synchronize_session=False)
        deleted_count = query.delete(synchronize_session=False)
        db.commit()
    else:
        deleted_count = 0
        
    return {"ok": True, "deleted_count": deleted_count, "message": f"{deleted_count} ta xodim o'chirildi"}


@router.post("/api/employees/preview-zip")
async def preview_employees_zip(
    request: Request,
    file: UploadFile = File(...),
    organization_id: int = Form(...),
    employee_type: str = Form("staff"),
    db: Session = Depends(get_db),
):
    import zipfile
    import io
    import openpyxl

    from routers.employees_parts.common import resolve_effective_org_id
    resolved_org_id = resolve_effective_org_id(request, db, int(organization_id))
    if not resolved_org_id:
        raise HTTPException(status_code=403, detail="Tashkilotga kirish ruxsati yo'q")

    zip_bytes = await file.read()
    try:
        z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fayl to'g'ri ZIP arxiv emas: {str(e)}")

    namelist = z.namelist()

    excel_file_path = None
    for name in namelist:
        if "__MACOSX" in name or name.split('/')[-1].startswith('.'):
            continue
        if name.lower().endswith(('.xlsx', '.xls')):
            excel_file_path = name
            break

    if not excel_file_path:
        raise HTTPException(status_code=400, detail="ZIP arxivi ichida Excel (.xlsx, .xls) fayli topilmadi")

    try:
        excel_data = z.read(excel_file_path)
        wb = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True)
        sheet = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel faylini o'qishda xatolik: {str(e)}")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {"ok": True, "items": []}

    headers = [str(cell).strip() if cell is not None else f"col_{idx}" for idx, cell in enumerate(rows[0])]
    items = []
    for r in rows[1:]:
        if all(cell is None for cell in r):
            continue
        row_dict = {}
        for idx, val in enumerate(r):
            if idx < len(headers):
                row_dict[headers[idx]] = val
        items.append(row_dict)

    preview_items = []

    for idx, item in enumerate(items):
        first_name = str(item.get("first_name") or item.get("Ism") or "").strip()
        last_name = str(item.get("last_name") or item.get("Familiya") or "").strip()
        if not first_name or not last_name:
            continue

        personal_id = str(item.get("personal_id") or item.get("Shaxsiy ID (Personal ID)") or "").strip()
        normalized_personal_id = normalize_personal_id(personal_id) if personal_id else None

        is_auto_id = not normalized_personal_id
        if is_auto_id:
            normalized_personal_id = generate_unique_personal_id(db)

        # Department
        dept_name = str(item.get("department") or item.get("Bo'lim") or item.get("Sinf / Guruh") or "").strip()

        # Position
        pos_name = str(item.get("position") or item.get("Lavozim") or "").strip()

        # Try to resolve photo from ZIP
        avatar_url = None
        img_filename = str(item.get("Rasm (URL yoki fayl nomi)") or item.get("Rasm") or item.get("image_url") or "").strip()
        print(f"[DEBUG IMPORT] Row #{idx+1} ({last_name} {first_name}): img_filename={repr(img_filename)}")
        if img_filename:
            base_name = img_filename.lower().split('/')[-1].split('\\')[-1]
            matched_in_zip = False
            for name in namelist:
                if "__MACOSX" in name or name.split('/')[-1].startswith('.'):
                    continue
                zip_base_name = name.lower().split('/')[-1].split('\\')[-1]
                if zip_base_name == base_name:
                    matched_in_zip = True
                    print(f"[DEBUG IMPORT] Match found in zip for base_name={repr(base_name)} -> zip entry={repr(name)}")
                    try:
                        image_bytes = z.read(name)
                        if image_bytes:
                            print(f"[DEBUG IMPORT] Read {len(image_bytes)} bytes from zip entry={repr(name)}")
                            avatar_url = _save_face_bytes_to_local_for_import(image_bytes)
                            print(f"[DEBUG IMPORT] _save_face_bytes_to_local_for_import returned: {repr(avatar_url)}")
                        else:
                            print(f"[DEBUG IMPORT] Empty bytes for zip entry={repr(name)}")
                    except Exception as _e:
                        print(f"[DEBUG IMPORT] Exception during image read/save: {str(_e)}")
                    break
            if not matched_in_zip:
                print(f"[DEBUG IMPORT] NO match in zip namelist for base_name={repr(base_name)}. List has {len(namelist)} items.")

        emp_type = str(item.get("employee_type") or item.get("Xodim turi (oqituvchi/hodim)") or item.get("O'quvchi turi (oquvchi/talaba)") or employee_type or "staff").strip().lower()
        if emp_type in ("oquvchi", "talaba", "students", "student"):
            emp_type = "student"
        else:
            emp_type = "staff"

        birth_date_str = ""
        raw_bdate_cell = item.get("birth_date") or item.get("Tug'ilgan sana (YYYY-MM-DD)")
        if isinstance(raw_bdate_cell, (datetime, date)):
            birth_date_str = raw_bdate_cell.strftime("%Y-%m-%d")
        elif raw_bdate_cell:
            birth_date_str = str(raw_bdate_cell).strip()

        phone = str(item.get("phone") or item.get("Telefon raqami") or "").strip() or None
        parent_phone = str(item.get("parent_phone") or item.get("Ota-onasining telefon raqami") or "").strip() or None
        region = str(item.get("region") or item.get("Viloyat") or "").strip() or None
        district = str(item.get("district") or item.get("Tuman") or "").strip() or None
        address = str(item.get("address") or item.get("Manzil") or "").strip() or None
        gender = str(item.get("gender") or item.get("Jinsi (male/female)") or "").strip().lower() or None

        preview_items.append({
            "first_name": first_name,
            "last_name": last_name,
            "middle_name": str(item.get("middle_name") or item.get("Otasining ismi") or "").strip() or None,
            "personal_id": normalized_personal_id,
            "is_auto_id": is_auto_id,
            "department": dept_name,
            "position": pos_name,
            "employee_type": emp_type,
            "birth_date": birth_date_str,
            "phone": phone,
            "parent_phone": parent_phone,
            "region": region,
            "district": district,
            "address": address,
            "gender": gender,
            "image_url": avatar_url,
        })

    return {"ok": True, "items": preview_items}


@router.post("/api/employees/confirm-import-zip")
def confirm_import_employees_zip(
    request: Request,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    organization_id = payload.get("organization_id")
    if not organization_id:
        raise HTTPException(status_code=422, detail="Tashkilot tanlanishi shart")

    from routers.employees_parts.common import resolve_effective_org_id
    resolved_org_id = resolve_effective_org_id(request, db, int(organization_id))
    if not resolved_org_id:
        raise HTTPException(status_code=403, detail="Tashkilotga kirish ruxsati yo'q")

    branch_id = payload.get("branch_id")
    if not branch_id:
        branches = db.query(Branch).filter(Branch.organization_id == resolved_org_id).order_by(Branch.id).all()
        if not branches:
            org_obj = db.query(Organization).filter(Organization.id == resolved_org_id).first()
            default_branch = Branch(
                organization_id=resolved_org_id,
                name="Asosiy filial",
                address=org_obj.address if org_obj else None,
                latitude=org_obj.latitude if org_obj else None,
                longitude=org_obj.longitude if org_obj else None,
                radius=(org_obj.radius or 100) if org_obj else 100,
            )
            db.add(default_branch)
            db.commit()
            db.refresh(default_branch)
            branch_id = default_branch.id
        else:
            branch_id = branches[0].id
    else:
        branch_id = int(branch_id)

    items = payload.get("items") or []
    if not items:
        return {"ok": True, "imported_count": 0, "message": "Yuklanadigan xodimlar mavjud emas"}

    # Pre-compute bcrypt hash once (bcrypt is slow, ~300ms per call)
    import bcrypt as _bcrypt
    _shared_hashed_pw = _bcrypt.hashpw(b'bioface', _bcrypt.gensalt()).decode('utf-8')

    dept_cache = {}
    pos_cache = {}
    imported_count = 0
    errors = []

    for idx, item in enumerate(items):
        try:
            first_name = str(item.get("first_name") or "").strip()
            last_name = str(item.get("last_name") or "").strip()
            if not first_name or not last_name:
                errors.append(f"Qator #{idx + 1}: Ism va Familiya bo'lishi shart")
                continue

            personal_id = str(item.get("personal_id") or "").strip()
            normalized_personal_id = normalize_personal_id(personal_id) if personal_id else None
            if not normalized_personal_id:
                normalized_personal_id = generate_unique_personal_id(db)
            else:
                dup = db.query(Employee).filter(
                    Employee.organization_id == resolved_org_id,
                    Employee.personal_id == normalized_personal_id
                ).first()
                if dup:
                    normalized_personal_id = generate_unique_personal_id(db)

            # Department
            dept_name = str(item.get("department") or "").strip()
            dept_id = None
            if dept_name:
                if dept_name not in dept_cache:
                    dept = db.query(Department).filter(
                        Department.organization_id == resolved_org_id,
                        func.lower(Department.name) == dept_name.lower()
                    ).first()
                    if not dept:
                        dept = Department(organization_id=resolved_org_id, name=dept_name)
                        db.add(dept)
                        db.commit()
                        db.refresh(dept)
                    dept_cache[dept_name] = dept.id
                dept_id = dept_cache[dept_name]

            # Position
            pos_name = str(item.get("position") or "").strip()
            pos_id = None
            if pos_name:
                if pos_name not in pos_cache:
                    pos = db.query(Position).filter(
                        Position.organization_id == resolved_org_id,
                        func.lower(Position.name) == pos_name.lower()
                    ).first()
                    if not pos:
                        pos = Position(organization_id=resolved_org_id, name=pos_name, department_id=dept_id)
                        db.add(pos)
                        db.commit()
                        db.refresh(pos)
                    pos_cache[pos_name] = pos.id
                pos_id = pos_cache[pos_name]

            emp_type = str(item.get("employee_type") or "staff").strip().lower()

            birth_date = None
            raw_bdate = str(item.get("birth_date") or "").strip()
            if raw_bdate:
                try:
                    birth_date = datetime.strptime(raw_bdate, "%Y-%m-%d")
                except:
                    try:
                        birth_date = datetime.strptime(raw_bdate, "%d.%m.%Y")
                    except:
                        pass

            emp = Employee(
                organization_id=resolved_org_id,
                branch_id=branch_id,
                first_name=first_name,
                last_name=last_name,
                middle_name=str(item.get("middle_name") or "").strip() or None,
                personal_id=normalized_personal_id,
                phone=str(item.get("phone") or "").strip() or None,
                parent_phone=str(item.get("parent_phone") or "").strip() or None,
                employee_type=emp_type,
                department_id=dept_id,
                position_id=pos_id,
                region=str(item.get("region") or "").strip() or None,
                district=str(item.get("district") or "").strip() or None,
                address=str(item.get("address") or "").strip() or None,
                birth_date=birth_date,
                gender=str(item.get("gender") or "").strip().lower() or None,
                image_url=item.get("image_url") or None,
                has_access=True,
            )
            db.add(emp)
            db.flush()  # get emp.id without full commit

            # Auto-create user account: username = personal_id, password = 'bioface'
            if emp.personal_id:
                try:
                    existing_user = db.query(User).filter(User.name == str(emp.personal_id)).first()
                    if not existing_user:
                        auto_user = User(
                            name=str(emp.personal_id),
                            email=f"{emp.personal_id}@bioface.local",
                            hashed_password=_shared_hashed_pw,
                            role=UserRole.tashkilot_admin,
                            status="active",
                            is_staff=False,
                            organization_id=resolved_org_id,
                            branch_id=branch_id,
                        )
                        db.add(auto_user)
                except Exception as _ue:
                    print(f"[AUTO USER] Failed to create user for employee {emp.personal_id}: {_ue}")

            imported_count += 1
        except Exception as e:
            db.rollback()
            errors.append(f"Qator #{idx + 1}: Kutilmagan xato: {str(e)}")
            continue

    # Single batch commit for all employees + users
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        errors.append(f"Batch commit xatosi: {str(e)}")

    # Trigger face embeddings after successful commit
    try:
        all_new_emps = db.query(Employee).filter(
            Employee.organization_id == resolved_org_id,
            Employee.image_url.isnot(None)
        ).order_by(Employee.id.desc()).limit(imported_count).all()
        for _emp in all_new_emps:
            trigger_embedding_generation_bg(employee_id=int(_emp.id))
    except Exception:
        pass

    return {"ok": True, "imported_count": imported_count, "errors": errors}

